"""
標籤模板引擎
支援兩種模板格式：
  1. 直接值格式（晶晟 Sheet1）：格子直接存 "料號：BAS089109BREB"
  2. 公式格式（晶晟 Sheet2）：A欄存公式 ="料號："&$E$2，E欄存實際值
  3. 橫向流水號格式（標籤.xlsx）：標題行 + 每行兩個流水號
"""
import copy
import json
import re
from datetime import datetime, timezone, timedelta

_TW = timezone(timedelta(hours=8))
from io import BytesIO

import openpyxl
from openpyxl.utils import get_column_letter


# ── 欄位對應表 ────────────────────────────────────────────────────
DYNAMIC_FIELDS = {
    "料號":       "item_no",
    "品名":       "name",
    "規格":       "description",
    "數量":       "quantity",
    "出貨日期":   "ship_date",
    "客戶料號":   "remark",
    "批號":       "lot_no",
    "銷貨單號":   "order_no",
    "客戶訂單":   "customer_order_no",
    "Lot No":    "lot_no",
    "流水號":     "sequence",
    "固定文字":   "__fixed__",
}

FIELD_LABELS = list(DYNAMIC_FIELDS.keys())


def _fmt_date(d: str) -> str:
    return d  # 直接回傳 YYYYMMDD，不加點


def _fmt_lot_no(order_no: str, seq: int) -> str:
    base = order_no[:8] if len(order_no) >= 8 else datetime.now(_TW).strftime("%Y%m%d")
    return f"{base}{seq:04d}"


def _get_value(item: dict, order: dict, field: str, seq: int = 1) -> str:
    if field in ("__fixed__", "", None):
        return None
    if field == "lot_no":
        lot = item.get("lot_no", "")
        return lot if lot else _fmt_lot_no(order.get("order_no", ""), seq)
    if field == "sequence":
        return f"{seq:04d}"
    if field == "ship_date":
        return _fmt_date(str(item.get("ship_date", "")))
    if field == "quantity":
        qty = item.get("quantity", "")
        unit = item.get("unit", "PCS")
        return f"{qty}{unit}" if qty else ""
    val = item.get(field) or order.get(field, "")
    return str(val) if val else ""


def _parse_formula_label(formula: str) -> str | None:
    """
    從公式 ="料號："&$E$2 抽出標籤前綴 "料號："
    回傳 None 代表不是這種公式
    """
    m = re.match(r'^="([^"]+)"&', formula)
    if m:
        return m.group(1)
    return None


def _guess_field(value: str) -> str:
    """根據格子的值或標籤文字猜測動態欄位"""
    if not value:
        return "__fixed__"
    v = value.lower()

    # 固定文字
    if "tel" in v or "fax" in v or "made in" in v:
        return "__fixed__"

    # inline 格式（"料號：xxx"）
    for zh, key in [
        ("料號", "item_no_inline"), ("品號", "item_no_inline"),
        ("品名", "name_inline"),
        ("規格", "description_inline"),
        ("數量", "quantity_inline"),
        ("出貨日期", "ship_date_inline"), ("出廠日期", "ship_date_inline"),
        ("批號", "lot_no_inline"),
        ("lot no", "lot_no_inline"), ("lot　no", "lot_no_inline"),
        ("客戶料號", "remark_inline"),
        ("po no", "customer_order_no_inline"), ("po　no", "customer_order_no_inline"),
    ]:
        if zh in v:
            return key

    # 標題欄位（橫向模式）
    if "流水碼" in v or "流水號" in v:  return "sequence"
    if "年月日" in v:                   return "ship_date"
    if "訂單號" in v:                   return "customer_order_no_inline"

    # 純值格
    if re.match(r'^\d{6,8}$', value):   return "ship_date"
    if re.match(r'^[A-Z]{2,}[0-9A-Z\-]+$', value) and len(value) >= 6:
        return "item_no"
    if re.match(r'^\d{4}$', value):     return "sequence"

    return "__fixed__"


def _guess_field_from_label(label: str) -> str:
    """從公式的前綴標籤（"料號："）猜欄位"""
    mapping = {
        "料號": "item_no_inline", "品號": "item_no_inline",
        "品名": "name_inline",
        "規格": "description_inline",
        "數量": "quantity_inline",
        "出貨日期": "ship_date_inline", "出廠日期": "ship_date_inline",
        "批號": "lot_no_inline",
        "lot no": "lot_no_inline", "lot　no": "lot_no_inline",
        "客戶料號": "remark_inline",
        "po no": "customer_order_no_inline", "po　no": "customer_order_no_inline",
        "銷貨單": "order_no_inline", "訂單": "customer_order_no_inline",
    }
    l = label.lower().rstrip("：:")
    for key, val in mapping.items():
        if key in l:
            return val
    return "__fixed__"


# ══════════════════════════════════════════════════════════════════
#  模板分析
# ══════════════════════════════════════════════════════════════════

def analyze_template(wb: openpyxl.Workbook, sheet_name: str) -> dict:
    ws = wb[sheet_name]

    # 找所有有值的格（含公式）
    cells_with_value = [
        cell for row in ws.iter_rows()
        for cell in row
        if cell.value is not None and str(cell.value).strip()
    ]
    if not cells_with_value:
        return {}

    max_row = max(c.row for c in cells_with_value)
    max_col = max(c.column for c in cells_with_value)

    # 偵測公式模式（A欄有公式、E欄有實際值）
    formula_cols = set()
    data_cols = set()
    for cell in cells_with_value:
        val = str(cell.value)
        if val.startswith('="') and '&' in val:
            formula_cols.add(cell.column)
        elif not val.startswith('='):
            data_cols.add(cell.column)

    is_formula_mode = bool(formula_cols)

    # 找分隔空欄（寬度<5 或全無值）
    col_has_value = set(c.column for c in cells_with_value)
    separator_cols = []
    for col_idx in range(1, max_col + 1):
        letter = get_column_letter(col_idx)
        width = ws.column_dimensions[letter].width if letter in ws.column_dimensions else 8
        if col_idx not in col_has_value or width < 5:
            separator_cols.append(col_idx)

    # 找並排單元的欄分組
    unit_col_groups = []
    current_group = []
    for col_idx in range(1, max_col + 1):
        if col_idx in separator_cols:
            if current_group:
                unit_col_groups.append(current_group)
                current_group = []
        else:
            current_group.append(col_idx)
    if current_group:
        unit_col_groups.append(current_group)

    # 公式模式：只取顯示欄（公式欄），資料欄（E欄）不計入單元
    if is_formula_mode:
        # 顯示欄 = 公式欄或固定文字欄（非資料值欄）
        display_col_groups = []
        for group in unit_col_groups:
            display_cols = [c for c in group if c not in data_cols or c in formula_cols]
            # 去掉只有資料值的欄（如 E欄）
            pure_data = [c for c in group if c in data_cols and c not in formula_cols]
            display = [c for c in group if c not in pure_data]
            if display:
                display_col_groups.append(display)
        if display_col_groups:
            unit_col_groups = display_col_groups

    units_per_row = len(unit_col_groups) if unit_col_groups else 1
    first_unit_cols = unit_col_groups[0] if unit_col_groups else [1]
    columns_per_unit = len(first_unit_cols)
    gap_cols = len([c for c in separator_cols if c < max(first_unit_cols[-1], 2)]) if units_per_row > 1 else 1

    # 偵測 unit_rows 和橫向模式
    header_rows = 0
    is_row_repeat_mode = False
    if not is_formula_mode and max_row >= 3:
        def strip_nums(s): return "".join(c for c in s if not c.isdigit())
        row2 = "".join(str(ws.cell(row=2, column=c).value or "") for c in first_unit_cols)
        row3 = "".join(str(ws.cell(row=3, column=c).value or "") for c in first_unit_cols)
        if strip_nums(row2) == strip_nums(row3) and row2 != row3:
            is_row_repeat_mode = True
            header_rows = 1
            unit_rows = 1

    if not is_row_repeat_mode:
        unit_rows = _detect_unit_rows(ws, first_unit_cols, max_row)

    # 擷取格子資訊
    cells_info = []

    def _add_cell(ws_cell, rel_row, rel_col, is_header=False):
        val = str(ws_cell.value).strip() if ws_cell.value else ""
        if not val:
            return

        # 判斷欄位類型
        formula_label = _parse_formula_label(val)
        if formula_label:
            field = _guess_field_from_label(formula_label)
            display_val = formula_label  # 只存標籤前綴
        else:
            field = _guess_field(val)
            if field.endswith("_inline"):
                # 只存前綴（"料號："），丟掉範本的樣品值（"BLSS400170122SA"）
                m = re.match(r'^(.*?[：:])[ \t]*', val)
                display_val = m.group(1) if m else val
            else:
                display_val = val

        cells_info.append({
            "row":       rel_row,
            "col":       rel_col,
            "value":     display_val,
            "field":     field,
            "is_header": is_header,
            "font_bold": ws_cell.font.bold if ws_cell.font else False,
            "alignment": ws_cell.alignment.horizontal if ws_cell.alignment else "left",
        })

    if is_row_repeat_mode:
        # 標題行
        for col_offset, col_idx in enumerate(first_unit_cols):
            _add_cell(ws.cell(row=1, column=col_idx), 1, col_offset + 1, is_header=True)
        # 資料行
        for col_offset, col_idx in enumerate(first_unit_cols):
            _add_cell(ws.cell(row=2, column=col_idx), 1, col_offset + 1, is_header=False)
    else:
        for row_idx in range(1, unit_rows + 1):
            for col_offset, col_idx in enumerate(first_unit_cols):
                cell = ws.cell(row=row_idx, column=col_idx)
                _add_cell(cell, row_idx, col_offset + 1)

    # 欄寬和行高
    col_widths = {}
    for i, col_idx in enumerate(first_unit_cols):
        letter = get_column_letter(col_idx)
        w = ws.column_dimensions[letter].width if letter in ws.column_dimensions else 10
        col_widths[str(i + 1)] = float(w)

    row_heights = {}
    for row_idx in range(1, unit_rows + (1 if is_row_repeat_mode else 0) + 1):
        h = ws.row_dimensions[row_idx].height if row_idx in ws.row_dimensions else 15
        row_heights[str(row_idx)] = float(h or 15)

    return {
        "sheet_name":           sheet_name,
        "unit_rows":            unit_rows,
        "header_rows":          header_rows,
        "is_row_repeat_mode":   is_row_repeat_mode,
        "is_formula_mode":      is_formula_mode,
        "is_passthrough":       True,
        "first_unit_start_col": first_unit_cols[0] if first_unit_cols else 1,
        "columns_per_unit":     columns_per_unit,
        "gap_cols":             max(gap_cols, 1),
        "units_per_row":        units_per_row,
        "cells":                cells_info,
        "col_widths":           col_widths,
        "row_heights":          row_heights,
    }


def _detect_unit_rows(ws, first_unit_cols: list[int], max_row: int) -> int:
    """
    找出模板的列重複週期（= 一個標籤單元的高度）。
    比對方式：將各欄位值去除數字後比較（樣本數值不同不影響結構比對）。
    若找不到週期，回傳 max_row（整張 sheet = 一個標籤）。
    最短週期設為 3，避免誤偵測。
    """
    _NUM = re.compile(r'[\d.,]+')

    def row_key(r: int) -> tuple:
        return tuple(
            _NUM.sub('', str(ws.cell(row=r, column=c).value or '')).strip()
            for c in first_unit_cols
        )

    keys = [row_key(r) for r in range(1, max_row + 1)]
    k0 = keys[0]
    if not any(k0):
        return max_row

    for p in range(3, max_row // 2 + 1):
        if keys[p] != k0:
            continue
        if all(keys[off] == keys[p + off]
               for off in range(p) if p + off < len(keys)):
            return p

    return max_row


def analyze_all_sheets(wb: openpyxl.Workbook) -> dict[str, dict]:
    """分析 workbook 中所有工作表，回傳 {sheet_name: template_info}"""
    results = {}
    for name in wb.sheetnames:
        try:
            info = analyze_template(wb, name)
            if info and info.get("cells"):
                results[name] = info
        except Exception:
            pass
    return results


# ══════════════════════════════════════════════════════════════════
#  標籤產生
# ══════════════════════════════════════════════════════════════════

def _fill_value(cell_info: dict, item: dict, order: dict, seq: int) -> str | None:
    """根據 cell_info 的 field 計算要填入的值"""
    field = cell_info.get("field", "__fixed__")
    label = cell_info.get("value", "")  # 前綴標籤或固定文字

    if field == "__fixed__":
        return label

    base_field = field.replace("_inline", "")

    if field.endswith("_inline"):
        # 格式：「標籤：值」合在一格
        new_val = _get_value(item, order, base_field, seq)
        # 從 label 中只取前綴（到第一個 ：或 : 為止），
        # 這樣即使 label 是「料號：BLSS400170122SA」也不會重複帶入範本值
        m = re.match(r'^(.*?[：:])[ \t]*', label)
        prefix = m.group(1) if m else (label.rstrip("：: ") + "：")
        return prefix + (new_val or "")
    else:
        return _get_value(item, order, field, seq)


def _write_passthrough_to_sheet(ws_out, ws_tmpl, template_info: dict, orders: list[dict]):
    """
    Passthrough 模式：直接複製 template 的格子（保留格式），再覆寫動態欄位。
    """
    unit_rows            = template_info["unit_rows"]
    columns_per_unit     = template_info["columns_per_unit"]
    units_per_row        = template_info.get("units_per_row", 1)
    gap_cols             = template_info.get("gap_cols", 1)
    cells_info           = template_info["cells"]
    row_heights          = template_info.get("row_heights", {})
    first_unit_start_col = template_info.get("first_unit_start_col", 1)

    data_cells = [c for c in cells_info
                  if not c.get("is_header") and c.get("field", "__fixed__") != "__fixed__"]

    # 欄寬：從 template 直接讀
    for c_off in range(columns_per_unit):
        tmpl_letter = get_column_letter(first_unit_start_col + c_off)
        width = ws_tmpl.column_dimensions[tmpl_letter].width \
                if tmpl_letter in ws_tmpl.column_dimensions else 10
        for ui in range(units_per_row):
            out_col = ui * (columns_per_unit + gap_cols) + c_off + 1
            ws_out.column_dimensions[get_column_letter(out_col)].width = width

    # 預計算 template 中在標籤範圍內的合併格
    unit_merges = [
        m for m in ws_tmpl.merged_cells.ranges
        if (1 <= m.min_row <= unit_rows and 1 <= m.max_row <= unit_rows and
            first_unit_start_col <= m.min_col < first_unit_start_col + columns_per_unit)
    ]

    all_items = [(item, order) for order in orders for item in order.get("items", [])]
    current_row = 1

    # 每列放不同品項（最多 units_per_row 張），不重複
    for group_start in range(0, len(all_items), units_per_row):
        group = all_items[group_start:group_start + units_per_row]

        # 行高
        for ri_str, h in row_heights.items():
            if h is not None:
                ws_out.row_dimensions[current_row + int(ri_str) - 1].height = h

        for ui, (item, order) in enumerate(group):
            seq = group_start + ui + 1
            base_col = ui * (columns_per_unit + gap_cols)

            # 1. 複製 template 所有格子（保留值 + 樣式）
            for r in range(1, unit_rows + 1):
                for c_off in range(columns_per_unit):
                    src = ws_tmpl.cell(row=r, column=first_unit_start_col + c_off)
                    dst = ws_out.cell(row=current_row + r - 1, column=base_col + c_off + 1)
                    if src.value is not None:
                        dst.value = src.value
                    try:
                        if src.font:
                            dst.font = copy.copy(src.font)
                        if src.fill and getattr(src.fill, 'fill_type', None) not in (None, 'none'):
                            dst.fill = copy.copy(src.fill)
                        if src.border:
                            dst.border = copy.copy(src.border)
                        if src.alignment:
                            dst.alignment = copy.copy(src.alignment)
                    except Exception:
                        pass

            # 2. 套用合併格
            for m in unit_merges:
                try:
                    ws_out.merge_cells(
                        start_row=current_row + m.min_row - 1,
                        start_column=base_col + m.min_col - first_unit_start_col + 1,
                        end_row=current_row + m.max_row - 1,
                        end_column=base_col + m.max_col - first_unit_start_col + 1,
                    )
                except Exception:
                    pass

            # 3. 覆寫動態格（用實際品項資料）
            for dc in data_cells:
                val = _fill_value(dc, item, order, seq)
                if val is not None:
                    ws_out.cell(row=current_row + dc["row"] - 1,
                                column=base_col + dc["col"]).value = val

        _copy_passthrough_images(ws_out, ws_tmpl, unit_rows, current_row)
        current_row += unit_rows + 1


def _write_order_to_sheet(
    ws_out,
    ws_tmpl,
    template_info: dict,
    orders: list[dict],
    logo_imgs: list = None,
):
    """Put all items from `orders` into `ws_out` using `template_info`."""
    is_row_repeat    = template_info.get("is_row_repeat_mode", False)

    if template_info.get("is_passthrough") and ws_tmpl and not is_row_repeat:
        _write_passthrough_to_sheet(ws_out, ws_tmpl, template_info, orders)
        return

    unit_rows        = template_info["unit_rows"]
    columns_per_unit = template_info["columns_per_unit"]
    gap_cols         = template_info.get("gap_cols", 1)
    units_per_row    = template_info.get("units_per_row", 2)
    cells_info       = template_info["cells"]
    col_widths       = template_info.get("col_widths", {})
    row_heights      = template_info.get("row_heights", {})

    header_cells = [c for c in cells_info if c.get("is_header")]
    data_cells   = [c for c in cells_info if not c.get("is_header")]

    # 設定欄寬（第 1 欄固定 24.5）
    for unit_idx in range(units_per_row):
        base = unit_idx * (columns_per_unit + gap_cols)
        for rel_col_str, width in col_widths.items():
            abs_col = base + int(rel_col_str)
            col_w = 24.5 if int(rel_col_str) == 1 else width
            ws_out.column_dimensions[get_column_letter(abs_col)].width = col_w

    all_items = [(item, order) for order in orders for item in order.get("items", [])]
    current_row = 1

    for item, order in all_items:
        qty = 1  # 一張標籤顯示整批數量，不重複印

        if is_row_repeat:
            # 每個 label 放置 logo（header row 起始）
            if logo_imgs:
                _place_label_images(ws_out, logo_imgs, current_row, ws_tmpl)

            if header_cells:
                for hc in header_cells:
                    cell = ws_out.cell(row=current_row, column=hc["col"], value=hc["value"])
                    _copy_style(ws_tmpl, 1, hc["col"], cell)
                if "1" in row_heights:
                    ws_out.row_dimensions[current_row].height = row_heights["1"]
                current_row += 1

            for row_offset in range(qty):
                seq_left  = row_offset * 2 + 1
                seq_right = row_offset * 2 + 2
                data_row  = current_row + row_offset
                for dc in data_cells:
                    col = dc["col"]
                    seq = seq_left if col <= columns_per_unit // 2 + 1 else seq_right
                    val = _fill_value(dc, item, order, seq)
                    cell = ws_out.cell(row=data_row, column=col, value=val)
                    _copy_style(ws_tmpl, 2, col, cell)
                ws_out.row_dimensions[data_row].height = row_heights.get("2", 20)

            current_row += qty + 1

        else:
            rows_needed = (qty + units_per_row - 1) // units_per_row
            for label_row in range(rows_needed):
                for rel_row_str, height in row_heights.items():
                    abs_row = current_row + int(rel_row_str) - 1
                    ws_out.row_dimensions[abs_row].height = height

                # 每個 label block 放置 logo
                if logo_imgs:
                    _place_label_images(ws_out, logo_imgs, current_row, ws_tmpl)

                for unit_idx in range(units_per_row):
                    label_seq = label_row * units_per_row + unit_idx + 1
                    if label_seq > qty:
                        break
                    base_col = unit_idx * (columns_per_unit + gap_cols)
                    for dc in data_cells:
                        abs_row = current_row + dc["row"] - 1
                        abs_col = base_col + dc["col"]
                        val = _fill_value(dc, item, order, label_seq)
                        out_cell = ws_out.cell(row=abs_row, column=abs_col, value=val)
                        if ws_tmpl:
                            _copy_style(ws_tmpl, dc["row"], dc["col"], out_cell)

                current_row += unit_rows

            current_row += 1


def generate_from_template(
    template_info: dict,
    orders: list[dict],
    template_wb: openpyxl.Workbook,
) -> BytesIO:
    """產出標籤 Excel：每張銷貨單一個工作表。"""
    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)  # 移除預設空白工作表

    sheet_name = template_info.get("sheet_name", "")
    ws_tmpl = template_wb[sheet_name] if sheet_name in template_wb.sheetnames else None

    for order in orders:
        ws_name = (order.get("order_no") or order.get("filename", "標籤"))[:31]
        # Excel 工作表名稱不可含 : \ / ? * [ ]
        for ch in r':\/? *[]':
            ws_name = ws_name.replace(ch, "_")
        ws_out = wb_out.create_sheet(title=ws_name)
        logo_imgs = _extract_logo_images(ws_tmpl) if ws_tmpl else []
        _write_order_to_sheet(ws_out, ws_tmpl, template_info, [order], logo_imgs=logo_imgs)

    if not wb_out.sheetnames:
        wb_out.create_sheet("出貨標籤")

    buf = BytesIO()
    wb_out.save(buf)
    buf.seek(0)
    return buf


def generate_labels_multiorder(
    order_template_pairs: list[dict],
) -> BytesIO:
    """
    每張銷貨單可指定不同模板，產出一個 Excel（每單一個工作表）。

    order_template_pairs: [
        {
            "order": <order_dict>,
            "template_info": <template_info_dict>,
            "template_wb": <openpyxl.Workbook>,  # 可為 None
        },
        ...
    ]
    """
    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)

    for pair in order_template_pairs:
        order = pair["order"]
        tinfo = pair["template_info"]
        twb   = pair.get("template_wb") or openpyxl.Workbook()

        sname = tinfo.get("sheet_name", "")
        ws_tmpl = twb[sname] if sname in twb.sheetnames else None

        ws_name = (order.get("order_no") or order.get("filename", "標籤"))[:31]
        for ch in r':\/? *[]':
            ws_name = ws_name.replace(ch, "_")
        ws_out = wb_out.create_sheet(title=ws_name)
        logo_imgs = _extract_logo_images(ws_tmpl) if ws_tmpl else []
        _write_order_to_sheet(ws_out, ws_tmpl, tinfo, [order], logo_imgs=logo_imgs)

    if not wb_out.sheetnames:
        wb_out.create_sheet("出貨標籤")

    buf = BytesIO()
    wb_out.save(buf)
    buf.seek(0)
    return buf


def _extract_logo_images(ws_src) -> list[dict]:
    """
    從模板提取 logo 圖片資訊。
    跳過 TwoCellAnchor 跨超過 3 列的圖片（背景/浮水印），只保留真正的 logo。
    回傳 list of {raw, w_px, h_px, col, rel_row (0-based)}
    """
    import io as _io
    result = []
    for img in getattr(ws_src, '_images', []):
        try:
            raw = None
            if hasattr(img, '_data') and callable(img._data):
                raw = img._data()
            elif hasattr(img, 'ref'):
                ref = img.ref
                if isinstance(ref, (bytes, bytearray)):
                    raw = bytes(ref)
                elif hasattr(ref, 'read'):
                    ref.seek(0)
                    raw = ref.read()
            if not raw:
                continue

            anchor  = img.anchor
            w_px = h_px = None
            rel_row   = 0     # 0-based row offset in template
            col_letter = "A"

            if hasattr(anchor, 'ext') and getattr(anchor.ext, 'cx', None):
                # OneCellAnchor — use EMU directly
                w_px = int(anchor.ext.cx / 914400 * 96)
                h_px = int(anchor.ext.cy / 914400 * 96)
                if hasattr(anchor, '_from'):
                    rel_row    = anchor._from.row
                    col_letter = get_column_letter(anchor._from.col + 1)
            elif hasattr(anchor, '_from') and hasattr(anchor, 'to'):
                # TwoCellAnchor：由 _copy_passthrough_images 直接複製，這裡跳過
                continue
            elif isinstance(anchor, str):
                m = re.match(r'^([A-Za-z]+)(\d+)$', anchor.strip())
                if m:
                    col_letter = m.group(1)
                    rel_row    = int(m.group(2)) - 1
                if img.width:
                    w_px, h_px = img.width, img.height

            if not w_px or not h_px:
                continue

            result.append({
                'raw': raw, 'w_px': w_px, 'h_px': h_px,
                'col': col_letter, 'rel_row': rel_row,
            })
        except Exception:
            pass

    # 每欄只保留最上面那個 logo（rel_row 最小），去掉重複/背景圖
    best: dict[str, dict] = {}
    for logo in result:
        col = logo['col']
        if col not in best or logo['rel_row'] < best[col]['rel_row']:
            best[col] = logo
    return list(best.values())


def _place_label_images(ws_out, logo_imgs: list, label_start_row: int, ws_tmpl=None):
    """每個 label 各放一份 logo，置中於所在儲存格。"""
    import io as _io
    from openpyxl.drawing.image import Image as XLImage

    for limg in logo_imgs:
        try:
            col     = limg['col']
            abs_row = label_start_row + limg['rel_row']  # 1-based
            w_px    = limg['w_px']
            h_px    = limg['h_px']

            new_img        = XLImage(_io.BytesIO(limg['raw']))
            new_img.width  = w_px
            new_img.height = h_px

            # 置中：用 OneCellAnchor + colOff/rowOff
            try:
                from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
                from openpyxl.drawing.xdr import XDRPositiveSize2D
                from openpyxl.utils import column_index_from_string

                col_idx = column_index_from_string(col) - 1   # 0-based
                row_idx = abs_row - 1                          # 0-based

                tmpl_r = limg['rel_row'] + 1
                row_h = (ws_tmpl.row_dimensions[tmpl_r].height or 15) if ws_tmpl else 15

                cell_h_emu = int(row_h * 12700)       # pt → EMU
                img_w_emu  = int(w_px * 9525)
                img_h_emu  = int(h_px * 9525)

                col_off = 0                                        # 靠左
                row_off = max(0, (cell_h_emu - img_h_emu) // 2)  # 上下置中

                anch        = OneCellAnchor()
                anch._from  = AnchorMarker(col=col_idx, colOff=col_off,
                                           row=row_idx, rowOff=row_off)
                anch.ext    = XDRPositiveSize2D(cx=img_w_emu, cy=img_h_emu)
                new_img.anchor = anch
            except Exception:
                new_img.anchor = f"{col}{abs_row}"

            ws_out.add_image(new_img)
        except Exception:
            pass


def _copy_passthrough_images(
    ws_out, ws_tmpl, unit_rows: int, label_start_row: int,
    col_offset: int = 0, only_cols: set = None
):
    """
    Copy TwoCellAnchor images from template to output with row/col offset.
    Preserves TwoCellAnchor structure so Excel renders correct image size.
    label_start_row: 1-indexed row in output where this label starts.
    col_offset: extra 0-based column offset (for multi-slot layouts).
    only_cols: if set, only copy images whose _from.col (0-based) is in this set.
    """
    import io as _io
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker

    seen: set = set()
    row_off = label_start_row - 1

    for img in getattr(ws_tmpl, '_images', []):
        anchor = img.anchor
        if not (hasattr(anchor, '_from') and hasattr(anchor, 'to')):
            continue
        fr = anchor._from.row
        fc = anchor._from.col
        if fr >= unit_rows:
            continue
        if only_cols is not None and fc not in only_cols:
            continue
        pos = (fr, fc)
        if pos in seen:
            continue
        seen.add(pos)
        try:
            # Read image bytes — try getvalue() first (BytesIO, position-independent)
            raw = None
            ref = getattr(img, 'ref', None)
            if ref is not None:
                if isinstance(ref, (bytes, bytearray)):
                    raw = bytes(ref)
                elif hasattr(ref, 'getvalue'):
                    raw = ref.getvalue()
                elif hasattr(ref, 'read'):
                    try:
                        ref.seek(0)
                    except Exception:
                        pass
                    raw = ref.read()
            if not raw:
                try:
                    d = img._data
                    raw = d() if callable(d) else (bytes(d) if d else None)
                except Exception:
                    pass
            if not raw:
                continue

            new_img = XLImage(_io.BytesIO(raw))
            # Explicitly construct anchor — avoids deepcopy failures on complex openpyxl objects
            new_anchor = TwoCellAnchor()
            new_anchor._from = AnchorMarker(
                col=fc + col_offset,
                colOff=getattr(anchor._from, 'colOff', 0),
                row=fr + row_off,
                rowOff=getattr(anchor._from, 'rowOff', 0),
            )
            new_anchor.to = AnchorMarker(
                col=anchor.to.col + col_offset,
                colOff=getattr(anchor.to, 'colOff', 0),
                row=anchor.to.row + row_off,
                rowOff=getattr(anchor.to, 'rowOff', 0),
            )
            try:
                new_anchor.editAs = anchor.editAs
            except Exception:
                pass
            new_img.anchor = new_anchor
            ws_out.add_image(new_img)
        except Exception:
            pass


def _copy_style(ws_tmpl, row: int, col: int, dst_cell):
    if ws_tmpl is None:
        return
    try:
        src = ws_tmpl.cell(row=row, column=col)
        if src.font:      dst_cell.font      = copy.copy(src.font)
        if src.fill and src.fill.fill_type != "none":
            dst_cell.fill = copy.copy(src.fill)
        if src.border:    dst_cell.border    = copy.copy(src.border)
        if src.alignment: dst_cell.alignment = copy.copy(src.alignment)
    except Exception:
        pass


# ── 序列化 ────────────────────────────────────────────────────────

def template_to_json(info: dict) -> str:
    keys = ["sheet_name","unit_rows","header_rows","is_row_repeat_mode",
            "is_formula_mode","is_passthrough","first_unit_start_col",
            "columns_per_unit","gap_cols","units_per_row",
            "cells","col_widths","row_heights"]
    return json.dumps({k: info[k] for k in keys if k in info}, ensure_ascii=False)


def template_from_json(s: str) -> dict:
    return json.loads(s)


def get_field_options() -> list[str]:
    return FIELD_LABELS


def write_lscr_labels(
    orders: list[dict],
    wb_tmpl,
    tmpl_info: dict,
    include_small: bool = True,
    include_large: bool = True,
) -> BytesIO:
    """
    LSCR 專用排版：小標籤兩張並排，大標籤置於第一列第三欄。
    每個品項：
      列 1：[小 0][小 1][大]
      列 2：[小 2][小 3]
      …
    """
    import math

    ws_tmpl = wb_tmpl["lable"]

    unit_rows            = tmpl_info["unit_rows"]
    columns_per_unit     = tmpl_info["columns_per_unit"]
    gap_cols             = tmpl_info.get("gap_cols", 1)
    first_unit_start_col = tmpl_info.get("first_unit_start_col", 1)
    data_cells = [c for c in tmpl_info["cells"]
                  if not c.get("is_header") and c.get("field", "__fixed__") != "__fixed__"]
    row_heights  = tmpl_info.get("row_heights", {})
    unit_width   = columns_per_unit + gap_cols  # 每個 slot 的欄數（含間距）

    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "Labels"

    # 欄寬：從 template 讀取（內容欄 + 間距欄），還原原始版面
    for ui in range(3):
        # 內容欄
        for c_off in range(columns_per_unit):
            ws_out.column_dimensions[get_column_letter(ui * unit_width + c_off + 1)].width = 25
        # 間距欄（slot 之間）
        for g in range(gap_cols):
            gap_tmpl_letter = get_column_letter(first_unit_start_col + columns_per_unit + g)
            gw = ws_tmpl.column_dimensions[gap_tmpl_letter].width \
                 if gap_tmpl_letter in ws_tmpl.column_dimensions else 3
            out_gap_col = ui * unit_width + columns_per_unit + g + 1
            ws_out.column_dimensions[get_column_letter(out_gap_col)].width = gw

    unit_merges = [
        m for m in ws_tmpl.merged_cells.ranges
        if (1 <= m.min_row <= unit_rows and 1 <= m.max_row <= unit_rows and
            first_unit_start_col <= m.min_col < first_unit_start_col + columns_per_unit)
    ]

    def _write_slot(item, order, seq, slot_idx, out_row):
        base_col = slot_idx * unit_width
        for r in range(1, unit_rows + 1):
            for c_off in range(columns_per_unit):
                src = ws_tmpl.cell(row=r, column=first_unit_start_col + c_off)
                dst = ws_out.cell(row=out_row + r - 1, column=base_col + c_off + 1)
                if src.value is not None:
                    dst.value = src.value
                try:
                    if src.font:
                        dst.font = copy.copy(src.font)
                    if src.fill and getattr(src.fill, 'fill_type', None) not in (None, 'none'):
                        dst.fill = copy.copy(src.fill)
                    if src.border:
                        dst.border = copy.copy(src.border)
                    if src.alignment:
                        dst.alignment = copy.copy(src.alignment)
                except Exception:
                    pass
        for m in unit_merges:
            try:
                ws_out.merge_cells(
                    start_row=out_row + m.min_row - 1,
                    start_column=base_col + m.min_col - first_unit_start_col + 1,
                    end_row=out_row + m.max_row - 1,
                    end_column=base_col + m.max_col - first_unit_start_col + 1,
                )
            except Exception:
                pass
        for dc in data_cells:
            val = _fill_value(dc, item, order, seq)
            if val is not None:
                ws_out.cell(row=out_row + dc["row"] - 1,
                            column=base_col + dc["col"]).value = val

    current_row = 1
    global_seq  = 1

    for order in orders:
        for phys_item in order.get("items", []):
            total   = float(phys_item.get("_total_qty") or phys_item.get("quantity") or 0)
            small_q = float(phys_item.get("_small_qty") or total)
            large_q = float(phys_item.get("_large_qty") or total)
            small_u = phys_item.get("_small_unit", "PCS")
            large_u = phys_item.get("_large_unit", "PCS")
            one_box = (small_q >= total or total == 0)

            smalls     = []
            large_item = None
            if one_box:
                if include_small or include_large:
                    smalls = [dict(phys_item)]
            else:
                if include_small:
                    for _ in range(2):
                        s = dict(phys_item)
                        s["quantity"] = str(int(small_q))
                        s["unit"]     = small_u
                        smalls.append(s)
                if include_large:
                    large_item = dict(phys_item)
                    large_item["quantity"] = str(int(large_q))
                    large_item["unit"]     = large_u

            n_small_rows = math.ceil(len(smalls) / 2) if smalls else 0
            n_rows       = max(n_small_rows, 1 if large_item else 0)

            for row_idx in range(n_rows):
                for ri_str, h in row_heights.items():
                    if h is not None:
                        ws_out.row_dimensions[current_row + int(ri_str) - 1].height = h + 2

                si0 = row_idx * 2
                if si0 < len(smalls):
                    _write_slot(smalls[si0], order, global_seq + si0, 0, current_row)

                si1 = row_idx * 2 + 1
                if si1 < len(smalls):
                    _write_slot(smalls[si1], order, global_seq + si1, 1, current_row)

                if row_idx == 0 and large_item:
                    _write_slot(large_item, order, global_seq + len(smalls), 2, current_row)

                # 複製 logo（slot 0 和 1：col A、col C）
                _copy_passthrough_images(ws_out, ws_tmpl, unit_rows, current_row)
                # 複製 logo（slot 2 大標籤：只取 col A 圖片，平移到第三欄）
                if row_idx == 0 and large_item:
                    _copy_passthrough_images(
                        ws_out, ws_tmpl, unit_rows, current_row,
                        col_offset=2 * unit_width, only_cols={0},
                    )

                current_row += unit_rows + 1

            global_seq += len(smalls) + (1 if large_item else 0)

    # ── 列印版面設定 ──────────────────────────────────────────────
    from openpyxl.worksheet.pagebreak import Break as _Break
    from openpyxl.worksheet.page import PageMargins as _PageMargins

    last_row = current_row - 1
    # 列印範圍：A1 到第三個 slot 的最後內容欄（col E = 2*unit_width+1）
    print_last_col = get_column_letter(2 * unit_width + columns_per_unit)
    ws_out.print_area = f"A1:{print_last_col}{last_row}"

    # 每個品項後插入分頁（每 unit_rows+1 列一頁）
    block = unit_rows + 1
    for br in range(block, last_row + 1, block):
        ws_out.row_breaks.append(_Break(id=br))

    ws_out.page_setup.orientation = 'landscape'
    ws_out.page_margins = _PageMargins(
        left=0.1, right=0.1, top=0.1, bottom=0.1, header=0, footer=0
    )
    ws_out.page_setup.fitToPage = True
    ws_out.page_setup.fitToWidth = 1
    ws_out.page_setup.fitToHeight = 0

    buf = BytesIO()
    wb_out.save(buf)
    buf.seek(0)
    return buf