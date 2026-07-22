"""
生產日報表彙總
把多個工作表（各站人員填寫的日報表）依「日期＋工單號碼＋料號＋站」彙總，
同一天、同一工單、同一料號、同一站若分散在多列／多個工作表，加總 OK／NG 數量。

欄位版面（日期／工單號碼／料號／發料數量／各站…）並非固定欄位編號，
而是逐表掃描第 4 列標題動態判斷，避免不同工作表欄位增減時解析錯位。
"""
import re
from copy import copy
from io import BytesIO

import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter

_HEADER_ROW = 4
_SUBHEADER_ROW = 5
_DATA_START_ROW = 6
_MAX_LEADING_SCAN = 10  # 標題列最多掃描幾欄來找日期/工單號碼/料號/發料數量

_LEADING_FIELD_MAP = {
    "日期": "date_col",
    "工單號碼": "workorder_col",
    "料號": "item_col",
    "發料數量": "issue_col",
}


def _cell(ws, row, col) -> str:
    v = ws.cell(row=row, column=col).value
    return "" if v is None else str(v).strip()


def _detect_layout(ws) -> dict:
    """
    掃描第 4 列標題，動態判斷「日期／工單號碼／料號／發料數量」欄位位置，
    以及站別（製程欄位群組）從第幾欄開始。找不到對應欄位時該值為 None。
    """
    layout = {"date_col": None, "workorder_col": None, "item_col": None,
              "issue_col": None, "station_start": None}
    for c in range(1, _MAX_LEADING_SCAN + 1):
        text = _cell(ws, _HEADER_ROW, c).split("\n")[0].strip()
        if not text:
            continue
        if text in _LEADING_FIELD_MAP:
            layout[_LEADING_FIELD_MAP[text]] = c
            continue
        layout["station_start"] = c
        break
    return layout


def _station_map(ws, station_start: int) -> list[dict]:
    """掃描第 4/5 列標題，找出各站的欄位範圍與 OK/NG/原因 欄位。"""
    max_col = ws.max_column
    starts = []
    for c in range(station_start, max_col + 1):
        v = _cell(ws, _HEADER_ROW, c)
        if v:
            starts.append((c, v.split("\n")[0].strip()))

    seen: dict[str, int] = {}
    stations = []
    for i, (c, name) in enumerate(starts):
        end = starts[i + 1][0] - 1 if i + 1 < len(starts) else max_col
        ok_col = ng_col = None
        reason_cols = []
        for cc in range(c, end + 1):
            sub = _cell(ws, _SUBHEADER_ROW, cc).split("\n")[0].strip()
            if sub == "OK":
                ok_col = cc
            elif sub == "NG":
                ng_col = cc
            elif sub == "原因":
                reason_cols.append(cc)

        # 同一份表若有重複站名（例如流程中出現兩次「外觀檢查」），加序號區分，避免彙總時誤合併
        seen[name] = seen.get(name, 0) + 1
        display_name = name if seen[name] == 1 else f"{name}{seen[name]}"

        stations.append({
            "name": display_name, "start": c,
            "ok_col": ok_col, "ng_col": ng_col, "reason_cols": reason_cols,
        })
    return stations


def _month_from_sheet(ws) -> int | None:
    m = re.match(r"(\d+)", _cell(ws, 2, 1))
    return int(m.group(1)) if m else None


def _num(ws, row, col) -> float:
    if col is None:
        return 0.0
    v = ws.cell(row=row, column=col).value
    if v is None or v == "":
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def parse_daily_report_workbook(wb: openpyxl.Workbook, summary_sheet: str = "總表") -> pd.DataFrame:
    """
    解析生產日報表活頁簿，回傳長表：欄位為 月, 日, 工單號碼, 料號, 站, OK, NG, 工作表。
    自動略過沒有「日期／料號」標題列的工作表（例如空白總表），
    有無「工單號碼」欄位皆可解析（沒有時該欄為空字串）。
    一律跳過「總表」本身：它是彙總後的結果，不是原始人員填寫的資料，
    若被當成輸入來源會跟它彙總來源的原始工作表重複加總。
    """
    records = []
    for sheet_name in wb.sheetnames:
        if sheet_name == summary_sheet:
            continue
        ws = wb[sheet_name]
        if ws.max_row < _DATA_START_ROW:
            continue
        layout = _detect_layout(ws)
        if layout["date_col"] is None or layout["item_col"] is None or layout["station_start"] is None:
            continue

        month = _month_from_sheet(ws)
        stations = _station_map(ws, layout["station_start"])

        for r in range(_DATA_START_ROW, ws.max_row + 1):
            item_no = _cell(ws, r, layout["item_col"])
            date_raw = _cell(ws, r, layout["date_col"])
            if not item_no or not date_raw:
                continue
            try:
                day = int(float(date_raw))
            except ValueError:
                continue
            workorder = _cell(ws, r, layout["workorder_col"]) if layout["workorder_col"] else ""

            for st_info in stations:
                ok = _num(ws, r, st_info["ok_col"])
                ng = _num(ws, r, st_info["ng_col"])
                if st_info["ok_col"] is None and st_info["ng_col"] is None:
                    # 沒有 OK/NG 細分（例如「包裝」），該站起始欄位的值視為數量
                    ok = _num(ws, r, st_info["start"])
                if ok == 0 and ng == 0:
                    continue
                records.append({
                    "月": month,
                    "日": day,
                    "工單號碼": workorder,
                    "料號": item_no,
                    "站": st_info["name"],
                    "OK": ok,
                    "NG": ng,
                    "工作表": sheet_name,
                })

    return pd.DataFrame(records, columns=["月", "日", "工單號碼", "料號", "站", "OK", "NG", "工作表"])


def aggregate_daily_report(df: pd.DataFrame) -> pd.DataFrame:
    """
    依「日期＋工單號碼＋料號＋站」彙總加總 OK／NG，回傳寬表：
    每列為一個 (月, 日, 工單號碼, 料號)，各站的 OK/NG 各佔一欄。
    """
    if df.empty:
        return pd.DataFrame()

    grouped = df.groupby(["月", "日", "工單號碼", "料號", "站"], as_index=False)[["OK", "NG"]].sum()
    station_order = list(dict.fromkeys(df["站"]))

    wide = grouped.pivot(index=["月", "日", "工單號碼", "料號"], columns="站", values=["OK", "NG"]).fillna(0)
    ordered_cols = [(kind, station) for station in station_order for kind in ("OK", "NG")]
    wide = wide.reindex(columns=pd.MultiIndex.from_tuples(ordered_cols))
    wide.columns = [f"{station}_{kind}" for kind, station in wide.columns]
    wide = wide.reset_index().sort_values(["月", "日", "工單號碼", "料號"]).reset_index(drop=True)
    for col in wide.columns[4:]:
        wide[col] = wide[col].astype(int)
    return wide


def _layout_usable(ws) -> bool:
    layout = _detect_layout(ws)
    return layout["date_col"] is not None and layout["item_col"] is not None and layout["station_start"] is not None


def build_summary_workbook(wb: openpyxl.Workbook, wide_df: pd.DataFrame,
                            summary_sheet: str = "總表") -> BytesIO:
    """
    把彙總結果填入活頁簿中的「總表」工作表（沿用原本表頭格式與樣式），
    並移到最前面；其餘工作表（各站原始資料）維持不動一併保留。

    若既有總表的版面缺少目前資料中的欄位（例如舊版總表沒有「工單號碼」欄，
    但各站表已新增該欄），改以目前版面最新的工作表複製表頭重建總表。
    """
    newest_template = next(
        (wb[n] for n in wb.sheetnames if n != summary_sheet and _layout_usable(wb[n])),
        None,
    )

    ws = wb[summary_sheet] if summary_sheet in wb.sheetnames else None
    if ws is not None and newest_template is not None:
        cur_layout = _detect_layout(ws)
        new_layout = _detect_layout(newest_template)
        if cur_layout.get("workorder_col") is None and new_layout.get("workorder_col") is not None:
            # 總表版面過舊（缺少工單號碼欄），改用最新版面重建
            wb.remove(ws)
            ws = None

    if ws is None:
        template_ws = newest_template
        if template_ws is None:
            raise ValueError("找不到可用的表頭範本，無法建立總表")
        ws = wb.copy_worksheet(template_ws)
        ws.title = summary_sheet

    layout = _detect_layout(ws)
    stations = _station_map(ws, layout["station_start"])

    # 清除總表既有資料列（若原本已殘留資料）
    if ws.max_row >= _DATA_START_ROW:
        for row in ws.iter_rows(min_row=_DATA_START_ROW, max_row=ws.max_row):
            for cell in row:
                cell.value = None

    # 範本通常只預先套好前面幾十列的格式（框線等），資料筆數超出範本列數時，
    # 多出來的列直接用 ws.cell() 寫值不會帶格式。這裡固定拿第一筆資料列（_DATA_START_ROW）
    # 當樣式範本，每一列都套用一次，確保超出範本原本列數的資料列也有一致格式。
    _max_col = ws.max_column
    _style_row = _DATA_START_ROW
    _row_height = ws.row_dimensions[_style_row].height
    _col_styles = {
        c: copy(ws.cell(row=_style_row, column=c)._style)
        for c in range(1, _max_col + 1)
    }

    for i, (_, r) in enumerate(wide_df.iterrows()):
        row_idx = _DATA_START_ROW + i
        if row_idx != _style_row:
            for c in range(1, _max_col + 1):
                ws.cell(row=row_idx, column=c)._style = copy(_col_styles[c])
            if _row_height is not None:
                ws.row_dimensions[row_idx].height = _row_height
        ws.cell(row=row_idx, column=layout["date_col"], value=int(r["日"]))
        if layout["workorder_col"] and r.get("工單號碼"):
            ws.cell(row=row_idx, column=layout["workorder_col"], value=r["工單號碼"])
        ws.cell(row=row_idx, column=layout["item_col"], value=r["料號"])
        for st_info in stations:
            ok_val = r.get(f"{st_info['name']}_OK", 0)
            ng_val = r.get(f"{st_info['name']}_NG", 0)
            if st_info["ok_col"] is None and st_info["ng_col"] is None:
                if ok_val:
                    ws.cell(row=row_idx, column=st_info["start"], value=int(ok_val))
                continue
            if st_info["ok_col"] is not None and ok_val:
                ws.cell(row=row_idx, column=st_info["ok_col"], value=int(ok_val))
            if st_info["ng_col"] is not None and ng_val:
                ws.cell(row=row_idx, column=st_info["ng_col"], value=int(ng_val))

    # 「原因」欄位不能整欄隱藏——原因代碼對照表（1~8）就疊在這些欄位最上方幾列，
    # 隱藏欄位會連對照表一起藏起來。總表複製自的範本可能已把這些欄位設為隱藏，
    # 這裡強制取消隱藏（欄位內容本來就是空的，顯示出來無妨）。
    reason_cols = sorted({c for st_info in stations for c in st_info["reason_cols"]})
    for c in reason_cols:
        ws.column_dimensions[get_column_letter(c)].hidden = False

    idx = wb.sheetnames.index(summary_sheet)
    if idx != 0:
        wb.move_sheet(summary_sheet, offset=-idx)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
