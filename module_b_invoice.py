"""
模組 B：銷貨單 → 電子發票上傳 Excel（.xls）
依照 e-invoice.com.tw V1.6 格式產生可直接上傳的 .xls 檔案。

規則：
- 銷貨單有發票號碼才匯入，沒有的直接跳過
- 發票人（賣方）統編固定 24405403
- 受票人（買方）統編從銷貨單的 buyer_tax_id 取
- 單價暫定 100（待使用者提供價目表後更新）
"""
from datetime import datetime
from io import BytesIO

import xlwt

SELLER_TAX_ID = "24405403"
DEFAULT_UNIT_PRICE = 100

# ── xlwt 樣式 ─────────────────────────────────────────────────────
_HDR = xlwt.easyxf(
    "font: bold on, colour white; "
    "pattern: pattern solid, fore_colour dark_blue; "
    "alignment: horiz centre, vert centre; "
    "borders: left thin, right thin, top thin, bottom thin"
)
_DAT = xlwt.easyxf(
    "alignment: horiz centre, vert centre; "
    "borders: left thin, right thin, top thin, bottom thin"
)
_DAT_L = xlwt.easyxf(
    "alignment: horiz left, vert centre; "
    "borders: left thin, right thin, top thin, bottom thin"
)


def _tw_date(date_str: str) -> str:
    """YYYYMMDD → 民國年 YYYMMDD（7 碼）"""
    if len(date_str) == 8:
        try:
            y = int(date_str[:4]) - 1911
            return f"{y}{date_str[4:]}"
        except ValueError:
            pass
    today = datetime.today()
    return f"{today.year - 1911}{today.month:02d}{today.day:02d}"


def _now_time() -> str:
    return datetime.now().strftime("%H%M%S")


def _col_width(chars: int) -> int:
    """xlwt 欄寬單位：1/256 個字元寬度"""
    return max(chars, 8) * 300


def generate_invoice_excel(
    orders: list[dict],
    seller_tax_id: str = SELLER_TAX_ID,
    invoice_prefix: str = "AA",   # 保留參數但已改用銷貨單發票號碼
    start_number: int = 1,
) -> BytesIO:
    """
    接收多張銷貨單，只處理有發票號碼的訂單，產生 .xls 格式上傳檔。
    回傳 BytesIO（.xls bytes）。
    """
    wb = xlwt.Workbook(encoding="utf-8")
    ws_main   = wb.add_sheet("發票主檔")
    ws_detail = wb.add_sheet("發票明細")

    # ── 主檔標題 ──────────────────────────────────────────────────
    main_headers = [
        "發票號碼(IVNO)", "發票日期(IVDAT)", "發票時間(IVTM)",
        "未稅金額(IVAMT)", "稅率別(TAXRID)", "營業稅額(SALTAXAMT)",
        "發票人統一編號(IVPESRFNO)", "受票人統一編號(TAIVPESRFNO)",
        "款項別(CAID)", "相關號碼(RELNO)", "原幣金額(OCRYAMT)",
        "匯率(EXR)", "幣別(CUCY)", "彙開(GROPMK)", "通關方式(CSTMMK)",
        "買方聯絡人(BUYRCTM)", "買方聯絡人部門(BUYRCTMDP)",
        "買受人電子郵件(CUEMAIL)", "總備註(COMT5)",
        "發票開立自動通知(OPNAUTNTI)", "作廢發票自動通知(CANCELAUTNTI)",
        "零稅率原因(ZEROTAXRATEREASON)",
    ]
    detail_headers = [
        "發票號碼(IVNO)", "項次(IT)", "品名(DSR)", "品名2(DSR2)",
        "數量(QTY1)", "單位(UN1)", "單價(UP)", "金額(AMT)",
        "相關號碼一(RELNO1)", "相關號碼二(RELNO2)",
    ]

    for col, h in enumerate(main_headers):
        ws_main.write(0, col, h, _HDR)
        ws_main.col(col).width = _col_width(len(h))

    for col, h in enumerate(detail_headers):
        ws_detail.write(0, col, h, _HDR)
        ws_detail.col(col).width = _col_width(len(h))

    # ── 填入資料 ──────────────────────────────────────────────────
    main_row   = 1   # xlwt 從 0 開始
    detail_row = 1

    for order in orders:
        # 沒有發票號碼的銷貨單直接略過
        inv_no = order.get("invoice_no", "").strip()
        if not inv_no:
            continue

        items = order.get("items", [])
        if not items:
            continue

        # 金額計算（單價暫定 100，未稅）
        total_amount = sum(
            item.get("quantity", 0) * (item.get("unit_price") or DEFAULT_UNIT_PRICE)
            for item in items
        )
        tax_amount = round(total_amount * 0.05)

        inv_date = _tw_date(order.get("order_date", ""))
        inv_time = _now_time()
        sid = seller_tax_id or SELLER_TAX_ID
        bid = order.get("buyer_tax_id", "")
        relno = order.get("order_no", "")

        main_vals = [
            inv_no, inv_date, inv_time,
            total_amount, 1, tax_amount,
            sid, bid,
            "Z", relno, "", 1, "TWD",
            "", "", "", "", "", "",
            "", "", "",
        ]
        for col, val in enumerate(main_vals):
            ws_main.write(main_row, col, val, _DAT)
        main_row += 1

        # 明細
        for idx, item in enumerate(items, 1):
            qty  = item.get("quantity", 1) or 1
            unit = item.get("unit", "PC")
            up   = item.get("unit_price") or DEFAULT_UNIT_PRICE
            amt  = qty * up
            # 品名 = 品名 + 規格 合一
            name = item.get("name", "")
            spec = item.get("description", "")
            desc = (name + " " + spec).strip() or item.get("item_no", "")
            desc2    = item.get("remark", "") or ""
            relno1   = item.get("item_no", "")

            det_vals = [
                inv_no, idx, desc, desc2 if desc2 else "",
                qty, unit, up, amt,
                relno1, "",
            ]
            for col, val in enumerate(det_vals):
                style = _DAT_L if col == 2 else _DAT
                ws_detail.write(detail_row, col, val, style)
            detail_row += 1

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── 月結驗收資訊 ─────────────────────────────────────────────────


_ACCEPTANCE_COLS = {
    "出貨單號": 1, "單號": 2, "品號": 3, "品名": 4, "規格": 5,
    "出貨數量": 9, "單價": 12, "幣別": 13, "金額(未稅)": 15,
}


def parse_acceptance_excel(content: bytes) -> list[dict]:
    """
    解析「驗收資訊」xlsx（月結用），回傳每行 dict：
      order_no, line_no, part_no, name, spec, qty, unit_price, amount, currency
    """
    import openpyxl as _xl
    wb = _xl.load_workbook(BytesIO(content), data_only=True)
    ws = wb.active

    def _s(r, c):
        return str(ws.cell(r, c).value or "").strip()

    def _n(r, c):
        try:
            return float(ws.cell(r, c).value or 0)
        except (TypeError, ValueError):
            return 0.0

    # 自動偵測欄位位置（比對第一列標頭）
    header_map = {}
    for c in range(1, ws.max_column + 1):
        h = str(ws.cell(1, c).value or "").strip()
        if h in _ACCEPTANCE_COLS:
            header_map[h] = c

    col = lambda key: header_map.get(key, _ACCEPTANCE_COLS[key])

    result = []
    for r in range(2, ws.max_row + 1):
        order_no = _s(r, col("出貨單號"))
        if not order_no:
            continue
        result.append({
            "order_no":   order_no,
            "line_no":    _s(r, col("單號")),
            "part_no":    _s(r, col("品號")),
            "name":       _s(r, col("品名")),
            "spec":       _s(r, col("規格")),
            "qty":        _n(r, col("出貨數量")),
            "unit_price": _n(r, col("單價")),
            "amount":     _n(r, col("金額(未稅)")),
            "currency":   _s(r, col("幣別")) or "NTD",
        })
    return result


def generate_invoice_from_acceptance(
    rows: list[dict],
    invoice_no: str,
    invoice_date: str,
    buyer_tax_id: str,
    seller_tax_id: str = SELLER_TAX_ID,
) -> BytesIO:
    """
    從 parse_acceptance_excel 的 rows 產生電子發票上傳 xls。
    invoice_date: YYYYMMDD 字串
    """
    wb = xlwt.Workbook(encoding="utf-8")
    ws_main   = wb.add_sheet("發票主檔")
    ws_detail = wb.add_sheet("發票明細")

    main_headers = [
        "發票號碼(IVNO)", "發票日期(IVDAT)", "發票時間(IVTM)",
        "未稅金額(IVAMT)", "稅率別(TAXRID)", "營業稅額(SALTAXAMT)",
        "發票人統一編號(IVPESRFNO)", "受票人統一編號(TAIVPESRFNO)",
        "款項別(CAID)", "相關號碼(RELNO)", "原幣金額(OCRYAMT)",
        "匯率(EXR)", "幣別(CUCY)", "彙開(GROPMK)", "通關方式(CSTMMK)",
        "買方聯絡人(BUYRCTM)", "買方聯絡人部門(BUYRCTMDP)",
        "買受人電子郵件(CUEMAIL)", "總備註(COMT5)",
        "發票開立自動通知(OPNAUTNTI)", "作廢發票自動通知(CANCELAUTNTI)",
        "零稅率原因(ZEROTAXRATEREASON)",
    ]
    detail_headers = [
        "發票號碼(IVNO)", "項次(IT)", "品名(DSR)", "品名2(DSR2)",
        "數量(QTY1)", "單位(UN1)", "單價(UP)", "金額(AMT)",
        "相關號碼一(RELNO1)", "相關號碼二(RELNO2)",
    ]

    for c, h in enumerate(main_headers):
        ws_main.write(0, c, h, _HDR)
        ws_main.col(c).width = _col_width(len(h))
    for c, h in enumerate(detail_headers):
        ws_detail.write(0, c, h, _HDR)
        ws_detail.col(c).width = _col_width(len(h))

    total_amount = sum(r.get("amount", 0) for r in rows)
    tax_amount   = round(total_amount * 0.05)

    main_vals = [
        invoice_no, _tw_date(invoice_date), _now_time(),
        total_amount, 1, tax_amount,
        seller_tax_id, buyer_tax_id,
        "Z", "", "", 1, "TWD",
        "", "", "", "", "", "",
        "", "", "",
    ]
    for c, val in enumerate(main_vals):
        ws_main.write(1, c, val, _DAT)

    for idx, row in enumerate(rows, 1):
        name = row.get("name", "")
        spec = row.get("spec", "")
        dsr  = (name + "　" + spec).strip() if spec else name
        det_vals = [
            invoice_no, idx,
            dsr, row.get("part_no", ""),
            row.get("qty", 0), "個",
            row.get("unit_price", 0), row.get("amount", 0),
            row.get("order_no", ""), row.get("line_no", ""),
        ]
        for c, val in enumerate(det_vals):
            style = _DAT_L if c == 2 else _DAT
            ws_detail.write(idx, c, val, style)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
