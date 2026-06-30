"""
模組 C：出貨標籤產生器
- 使用者可勾選要顯示的欄位
- 可調整欄位順序
- 自動記住每家廠商的設定（透過 sheets_db）
"""
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

THIN        = Side(style="thin",   color="000000")
FULL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT        = Alignment(horizontal="left",   vertical="center", wrap_text=True)
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=12)
ALT_FILL    = PatternFill("solid", fgColor="EBF3FB")
LABEL_FONT  = Font(bold=True, size=9)

# 所有可選欄位：顯示名稱 → 取值 function
ALL_FIELDS = {
    "料號":     lambda item, order: item.get("item_no", ""),
    "品名":     lambda item, order: item.get("description", ""),
    "數量":     lambda item, order: str(item.get("quantity", "")),
    "單位":     lambda item, order: item.get("unit", "PC"),
    "客戶料號": lambda item, order: item.get("remark", ""),
    "批號":     lambda item, order: item.get("lot_no", ""),
    "客戶名稱": lambda item, order: item.get("customer", ""),
    "出貨日期": lambda item, order: _fmt_date(item.get("ship_date", "")),
    "銷貨單號": lambda item, order: order.get("order_no", ""),
    "客戶訂單": lambda item, order: order.get("customer_order_no", ""),
    "賣方統編": lambda item, order: order.get("seller_tax_id", ""),
    "買方統編": lambda item, order: order.get("buyer_tax_id", ""),
    "聯絡人":   lambda item, order: order.get("contact", ""),
}

DEFAULT_FIELDS = ["料號", "品名", "數量", "客戶料號", "批號", "客戶名稱", "出貨日期", "銷貨單號"]


def _fmt_date(d: str) -> str:
    if len(d) == 8:
        return f"{d[:4]}/{d[4:6]}/{d[6:8]}"
    return d


def _draw_label(ws, row_start: int, item: dict, order: dict, fields: list[str]) -> int:
    r = row_start
    customer = item.get("customer") or order.get("order_no", "出貨標籤")

    # 標題
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    cell = ws.cell(row=r, column=1, value=customer)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = CENTER
    cell.border = FULL_BORDER
    r += 1

    for field in fields:
        getter = ALL_FIELDS.get(field)
        if not getter:
            continue
        value = getter(item, order)
        if not value:
            continue
        # 欄位名稱
        lc = ws.cell(row=r, column=1, value=field)
        lc.font = LABEL_FONT
        lc.fill = ALT_FILL
        lc.alignment = CENTER
        lc.border = FULL_BORDER
        # 值（合併 2-4 欄）
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        vc = ws.cell(row=r, column=2, value=value)
        vc.font = Font(bold=(field in ["料號", "數量"]), size=10)
        vc.alignment = LEFT
        vc.border = FULL_BORDER
        r += 1

    r += 1
    return r


def generate_labels_excel(orders: list[dict], fields: list[str] | None = None) -> BytesIO:
    if fields is None:
        fields = DEFAULT_FIELDS
    wb = Workbook()
    ws = wb.active
    ws.title = "出貨標籤"
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10

    current_row = 1
    for order in orders:
        for item in order.get("items", []):
            current_row = _draw_label(ws, current_row, item, order, fields)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def get_all_field_names() -> list[str]:
    return list(ALL_FIELDS.keys())
