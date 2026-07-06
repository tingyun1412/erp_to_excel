"""
LSCR 出貨明細確認單 Excel 解析器
工作表 'list' → order/item 格式（與 RTF 解析器相容）
"""
import re
import openpyxl


def _v(ws, row, col) -> str:
    v = ws.cell(row=row, column=col).value
    return str(v).strip() if v is not None else ""


def parse_lscr_excel_wb(wb: openpyxl.Workbook) -> list[dict]:
    """
    從已開啟的 Workbook 解析 'list' 工作表。
    回傳 list[order_dict]，每個 PO NO 一筆。

    每個 item 含標準欄位外，另有私有欄位：
      _total_qty  : 總出貨數量 (str)
      _large_qty  : 大包裝數量 (str)
      _large_unit : 大包裝單位
      _small_qty  : 小包裝數量 (str)
      _small_unit : 小包裝單位
    """
    if "list" not in wb.sheetnames:
        raise ValueError("找不到 'list' 工作表")
    ws = wb["list"]

    # ── 標頭資訊（前 5 行）──────────────────────────────────────────
    customer = ""
    ship_date = ""
    for r in range(1, 6):
        for c in range(1, 20):
            v = _v(ws, r, c)
            if re.search(r'客戶[：:]', v):
                customer = re.sub(r'^.*客戶[：:]\s*', '', v).strip()
            if re.search(r'出貨日期[：:]', v):
                d = re.sub(r'^.*出貨日期[：:]\s*', '', v).strip()
                ship_date = re.sub(r'[-/年月]', '', d).strip('日').replace(' ', '')

    # ── 資料行
    # 欄位：C1=ID1, C2=ID2,
    #       C5=PO NO, C6=LOT NO, C7=Item(料號), C8=成品圖號,
    #       C9=品名, C10=規格,
    #       C11=總數量, C12=大包裝qty, C13=大包裝unit,
    #       C14=小包裝qty, C15=小包裝unit
    orders_map: dict[str, dict] = {}

    for r in range(6, ws.max_row + 1):
        item_col = _v(ws, r, 7)    # Item (料號)
        total_col = _v(ws, r, 11)  # 總數量
        if not item_col or not total_col:
            continue
        if item_col.lower() in ("item", "料號", "品名", "no."):
            continue

        id1 = _v(ws, r, 1)
        id2 = _v(ws, r, 2)
        po_no = _v(ws, r, 5) or "N/A"
        lot_no = _v(ws, r, 6)
        remark = _v(ws, r, 8)   # 成品圖號
        name = _v(ws, r, 9)
        desc = _v(ws, r, 10)
        large_qty = _v(ws, r, 12)
        large_unit = _v(ws, r, 13) or "PCS"
        small_qty = _v(ws, r, 14)
        small_unit = _v(ws, r, 15) or "PCS"

        # 品名加入 ID1/ID2 尺寸
        if id1 and id2:
            name = f"{name}（ID1={id1}mm*ID2={id2}mm）"

        # 預設用小包裝數量顯示
        qty = small_qty if small_qty else total_col
        unit = small_unit if small_qty else large_unit

        item = {
            "item_no":      item_col,
            "name":         name,
            "description":  desc,
            "quantity":     qty,
            "unit":         unit,
            "lot_no":       lot_no,
            "remark":       remark,
            "ship_date":    ship_date,
            # 私有欄位供展開邏輯使用
            "_total_qty":   total_col,
            "_large_qty":   large_qty or total_col,
            "_large_unit":  large_unit,
            "_small_qty":   small_qty,
            "_small_unit":  small_unit,
        }

        if po_no not in orders_map:
            orders_map[po_no] = {
                "order_no":          po_no,
                "customer_name":     customer,
                "customer_order_no": po_no,
                "ship_date":         ship_date,
                "items":             [],
            }
        orders_map[po_no]["items"].append(item)

    return list(orders_map.values())


def expand_lscr_items(orders: list[dict],
                      include_small: bool = True,
                      include_large: bool = True) -> list[dict]:
    """
    依大/小包裝展開品項：
    - include_small: 產生 N 張小包裝標籤（每張 = 小包裝數量）
    - include_large: 產生 1 張大包裝標籤（= 總出貨數量）
    若大小包裝數量相同（只有一箱），只印一張。
    """
    import math
    result = []
    for o in orders:
        new_o = dict(o)
        new_items = []
        for itm in o.get("items", []):
            total = float(itm.get("_total_qty") or itm.get("quantity") or 0)
            small = float(itm.get("_small_qty") or total)
            large = float(itm.get("_large_qty") or total)
            large_unit = itm.get("_large_unit") or itm.get("unit") or "PCS"
            small_unit = itm.get("_small_unit") or itm.get("unit") or "PCS"

            one_box = (small >= total or total == 0)

            if one_box:
                # 只有一箱，印一張
                new_items.append(dict(itm))
            else:
                if include_small and small > 0:
                    n = math.ceil(total / small)
                    for _ in range(n):
                        s = dict(itm)
                        s["quantity"] = str(int(small))
                        s["unit"] = small_unit
                        new_items.append(s)
                if include_large:
                    l = dict(itm)
                    l["quantity"] = str(int(large))
                    l["unit"] = large_unit
                    new_items.append(l)

        new_o["items"] = new_items
        result.append(new_o)
    return result


def parse_lscr_excel(path: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    return parse_lscr_excel_wb(wb)
