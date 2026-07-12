"""
出貨自動化工具 v3
模組：
  A - 電子發票產生
  B - 標籤模板管理（上傳舊標籤 → 設定欄位 → 產出新標籤）
"""
import json
import re
import tempfile
from io import BytesIO

import pandas as pd
import streamlit as st
import openpyxl

from rtf_parser import parse_sales_order_rtf
from lscr_parser import parse_lscr_excel_wb
from module_b_invoice import (
    generate_invoice_excel,
    parse_acceptance_excel,
    generate_invoice_from_acceptance,
)
from template_engine import (
    analyze_template, analyze_all_sheets,
    generate_from_template, generate_labels_multiorder,
    template_to_json, template_from_json,
    get_field_options, FIELD_LABELS, DYNAMIC_FIELDS,
    write_lscr_labels,
)
from sheets_db import (
    load_templates, save_template, delete_template,
    clear_cache,
    load_vendors, save_vendor, delete_vendor,
    download_template_excel,
    find_lscr_base_template_id, save_lscr_base_template, download_lscr_base_template,
)

st.set_page_config(page_title="出貨自動化工具", page_icon="📦", layout="wide")
st.title("📦 出貨自動化工具")


def _tmpl_label(r: dict) -> str:
    """模板顯示名稱：廠商=模板名稱時只顯示一個，否則顯示『廠商 — 模板』"""
    v, t = r.get("廠商名稱", ""), r.get("模板名稱", "")
    return v if v == t else f"{v} — {t}"

if "parsed_orders" not in st.session_state:
    st.session_state.parsed_orders = []
if "template_wb_bytes" not in st.session_state:
    st.session_state.template_wb_bytes = {}  # {template_key: bytes}



# ════════════════════════════════════════════════════════════════
#  側邊欄：上傳銷貨單
# ════════════════════════════════════════════════════════════════
with st.sidebar:
    st.header("上傳銷貨單")
    uploaded_files = st.file_uploader(
        "選擇 RTF 銷貨單（可多選）",
        type=["rtf"],
        accept_multiple_files=True,
    )

    if uploaded_files:
        if st.button("解析並匯入", type="primary", use_container_width=True):
            orders, errors = [], []
            with st.spinner("解析中..."):
                for uf in uploaded_files:
                    with tempfile.NamedTemporaryFile(suffix=".rtf", delete=False) as tmp:
                        tmp.write(uf.read())
                        tmp_path = tmp.name
                    try:
                        data = parse_sales_order_rtf(tmp_path)
                        data["filename"] = uf.name
                        orders.append(data)
                    except Exception as e:
                        errors.append(f"{uf.name}: {e}")

            for err in errors:
                st.error(err)

            if orders:
                st.session_state.parsed_orders = orders
                st.success(f"解析完成，共 {len(orders)} 張銷貨單")
                st.rerun()

    st.divider()
    if st.session_state.parsed_orders:
        st.success(f"已載入 {len(st.session_state.parsed_orders)} 張銷貨單")
        if st.button("清除暫存", use_container_width=True):
            st.session_state.parsed_orders = []
            st.rerun()


# ════════════════════════════════════════════════════════════════
#  主頁籤
# ════════════════════════════════════════════════════════════════
tab_label, tab_invoice = st.tabs([
    "🏷 出貨標籤",
    "🧾 電子發票",
])


# ════════════════════════════════════════════════════════════════
#  標籤模板管理
# ════════════════════════════════════════════════════════════════
with tab_label:
    st.subheader("出貨標籤")

    # 頁面一開啟就從 Google Drive 補載所有有 ID 的模板 Excel（不需等有訂單才跑）
    _preload_missing = []   # 雲端根本沒有 Excel檔案ID（之前上傳時就沒同步成功）
    _preload_failed  = []   # 有 Excel檔案ID 但這次下載失敗（網路/權限問題）
    try:
        _all_tpls_preload = load_templates()
        for _tr in _all_tpls_preload:
            _tk = f"{_tr['廠商名稱']}_{_tr['模板名稱']}"
            if _tk in st.session_state.template_wb_bytes:
                continue
            if not _tr.get("Excel檔案ID"):
                _preload_missing.append(_tmpl_label(_tr))
                continue
            try:
                st.session_state.template_wb_bytes[_tk] = download_template_excel(_tr["Excel檔案ID"])
            except Exception:
                _preload_failed.append(_tmpl_label(_tr))
    except Exception:
        pass

    if _preload_missing or _preload_failed:
        with st.expander(f"⚠️ {len(_preload_missing) + len(_preload_failed)} 個模板需要重新上傳原始 Excel", expanded=False):
            if _preload_missing:
                st.caption("尚未同步到雲端（請到「管理模板」重新分析補傳一次即可永久解決）：" + "、".join(_preload_missing))
            if _preload_failed:
                st.caption("雲端下載失敗（可能是暫時性網路問題，重新整理再試一次）：" + "、".join(_preload_failed))

    sub_tab_use, sub_tab_manage, sub_tab_erp, sub_tab_lscr = st.tabs(["產出標籤", "管理模板", "從廠商網站下載標籤", "LSCR 確認單"])

    # ── 產出標籤 ──────────────────────────────────────────────
    with sub_tab_use:
        all_orders = st.session_state.parsed_orders

        if not all_orders:
            st.info("請先在左側上傳並解析銷貨單")
        else:
            try:
                all_templates = load_templates()
            except Exception as e:
                st.error(f"載入模板失敗：{e}")
                all_templates = []

            if not all_templates:
                st.warning("尚無任何標籤模板，請先到「管理模板」上傳")
            else:
                def _match_template(customer_name: str):
                    """廠商名稱比對：先完全包含，再找最長公共子字串 ≥ 3 字"""
                    if not customer_name:
                        return None

                    def _lcs_len(a: str, b: str) -> int:
                        best = 0
                        for i in range(len(a)):
                            for j in range(len(b)):
                                l = 0
                                while i + l < len(a) and j + l < len(b) and a[i + l] == b[j + l]:
                                    l += 1
                                if l > best:
                                    best = l
                        return best

                    # 1. 優先：完全包含（原邏輯）
                    for r in all_templates:
                        vendor = r.get("廠商名稱", "")
                        if vendor and (vendor in customer_name or customer_name in vendor):
                            return r

                    # 2. 次要：最長公共子字串 ≥ 3 字
                    best_rec, best_len = None, 2
                    for r in all_templates:
                        vendor = r.get("廠商名稱", "")
                        if vendor:
                            ln = _lcs_len(customer_name, vendor)
                            if ln > best_len:
                                best_len, best_rec = ln, r
                    return best_rec

                # 選銷貨單
                order_options = {
                    o.get("order_no", o.get("filename", f"單{i}")): o
                    for i, o in enumerate(all_orders)
                }
                selected_order_nos = st.multiselect(
                    "選擇要產生標籤的銷貨單",
                    options=list(order_options.keys()),
                    default=list(order_options.keys()),
                )
                selected_orders = [order_options[k] for k in selected_order_nos]

                template_options = [
                    _tmpl_label(r)
                    for r in all_templates
                ]

                # 每張銷貨單獨立選擇模板（自動比對可手動覆蓋）
                order_tmpl_map = {}
                if selected_orders:
                    st.markdown("**各銷貨單模板（自動比對，可手動覆蓋）：**")
                    for o in selected_orders:
                        order_key = o.get("order_no", o.get("filename", ""))
                        cname = o.get("customer_name", "")
                        auto_match = _match_template(cname)
                        default_idx = next(
                            (i for i, r in enumerate(all_templates) if r is auto_match), 0
                        )
                        _c1, _c2 = st.columns([5, 5])
                        with _c1:
                            _lbl = "✅ 自動" if auto_match else "⚠️ 未比對"
                            st.caption(f"{_lbl}　{order_key}　{cname or '未知客戶'}")
                        with _c2:
                            _chosen = st.selectbox(
                                "模板",
                                options=range(len(template_options)),
                                format_func=lambda i: template_options[i],
                                index=default_idx,
                                key=f"tmpl_sel_{order_key}",
                                label_visibility="collapsed",
                            )
                        order_tmpl_map[order_key] = all_templates[_chosen]

                # 顯示選取的品項
                selected_items = [
                    (item, order)
                    for order in selected_orders
                    for item in order.get("items", [])
                ]
                if selected_items:
                    st.dataframe(
                        [{
                            "銷貨單號": order.get("order_no",""),
                            "料號":     item.get("item_no",""),
                            "品名":     item.get("name",""),
                            "規格":     item.get("description",""),
                            "數量":     item.get("quantity",""),
                            "客戶料號": item.get("remark",""),
                            "批號":     item.get("lot_no",""),
                        } for item, order in selected_items],
                        use_container_width=True,
                        hide_index=True,
                        height=min(420, 38 * (len(selected_items) + 1) + 10),
                        column_config={
                            "銷貨單號": st.column_config.TextColumn(width="medium"),
                            "料號":     st.column_config.TextColumn(width="large"),
                            "品名":     st.column_config.TextColumn(width="small"),
                            "規格":     st.column_config.TextColumn(width="large"),
                            "數量":     st.column_config.NumberColumn(width="small"),
                            "客戶料號": st.column_config.TextColumn(width="medium"),
                            "批號":     st.column_config.TextColumn(width="medium"),
                        },
                    )
                    st.caption(f"共 {len(selected_items)} 個品項，每張銷貨單一個工作表")

                    # ── 分裝設定 ─────────────────────────────────────────
                    with st.expander("📦 分裝設定"):
                        _pkg_enable = st.checkbox("啟用分裝", key="pkg_enable")
                        if _pkg_enable:
                            st.caption(
                                "填入每箱數量；若每箱數量 ≥ 總數量表示只有一箱，只印一張標籤。"
                                "否則固定印出 2 張小標籤（顯示每箱數量）+ 1 張大標籤（顯示總數量）並排，"
                                "不會依箱數自動算張數——實際要印幾份由印標籤的人自己決定。"
                            )
                            _pkg_df = pd.DataFrame([
                                {
                                    "料號":   itm.get("item_no", ""),
                                    "品名":   itm.get("name", ""),
                                    "總數量": int(float(itm.get("quantity") or 0)),
                                    "每箱數量": int(float(itm.get("quantity") or 1)),
                                    "小標籤": True,
                                    "大標籤": True,
                                }
                                for itm, _ in selected_items
                            ])
                            _pkg_edited = st.data_editor(
                                _pkg_df,
                                column_config={
                                    "料號":   st.column_config.TextColumn(disabled=True, width="medium"),
                                    "品名":   st.column_config.TextColumn(disabled=True, width="small"),
                                    "總數量": st.column_config.NumberColumn(disabled=True, width="small"),
                                    "每箱數量": st.column_config.NumberColumn(min_value=1, width="small"),
                                    "小標籤": st.column_config.CheckboxColumn(width="small"),
                                    "大標籤": st.column_config.CheckboxColumn(width="small"),
                                },
                                hide_index=True,
                                use_container_width=True,
                                key="pkg_table",
                            )
                        else:
                            _pkg_edited = None

                else:
                    _pkg_enable = False
                    _pkg_edited = None
                    st.warning("請選擇至少一張銷貨單")

                # 警告：模板無動態欄位 → 標籤只有固定文字
                if selected_orders and order_tmpl_map:
                    for _rec in order_tmpl_map.values():
                        _info = template_from_json(_rec["設定JSON"])
                        _dyn = [c for c in _info.get("cells", []) if c.get("field") != "__fixed__"]
                        if not _dyn:
                            st.error(
                                f"⚠️ 模板「{_tmpl_label(_rec)}」沒有動態欄位，"
                                "標籤將只顯示固定文字（無料號、品名等）。"
                                "請到「管理模板」重新上傳並分析此模板。"
                            )
                            break

                if selected_orders and order_tmpl_map and st.button("產出標籤 Excel", type="primary", use_container_width=True):
                    with st.spinner("產出中..."):
                        try:
                            # 套用分裝：展開品項
                            def _expand_orders(orders, pkg_df):
                                result = []
                                _pkg_gid = [0]
                                for o in orders:
                                    new_o = dict(o)
                                    new_items = []
                                    for itm in o.get("items", []):
                                        total = float(itm.get("quantity") or 0)
                                        rows = pkg_df[pkg_df["料號"] == itm.get("item_no", "")]
                                        if rows.empty:
                                            new_items.append(dict(itm))
                                            continue
                                        row = rows.iloc[0]
                                        pkg = max(1.0, float(row["每箱數量"] or total))
                                        use_small = bool(row["小標籤"])
                                        use_large = bool(row["大標籤"])
                                        if pkg >= total or total == 0:
                                            # 只有一箱，印一張
                                            new_items.append(dict(itm))
                                        else:
                                            # 固定 2 張小標籤（每箱數量）+ 1 張大標籤（總數量）並排在同一列，
                                            # 不論每箱數量填多少都印 3 張——箱數由印標籤的人自己決定。
                                            # _pkg_group 標記讓產出時強制排在同一列，不受範本並排數限制。
                                            _pkg_gid[0] += 1
                                            gid = _pkg_gid[0]
                                            if use_small:
                                                s = dict(itm)
                                                s["quantity"] = int(pkg)
                                                s["_pkg_group"] = gid
                                                new_items.append(s)
                                                new_items.append(dict(s))
                                            if use_large:
                                                l = dict(itm)
                                                l["_pkg_group"] = gid
                                                new_items.append(l)
                                    new_o["items"] = new_items
                                    result.append(new_o)
                                return result

                            _gen_orders = (
                                _expand_orders(selected_orders, _pkg_edited)
                                if _pkg_enable and _pkg_edited is not None
                                else selected_orders
                            )

                            pairs = []
                            for o in _gen_orders:
                                order_key = o.get("order_no", o.get("filename", ""))
                                rec = order_tmpl_map.get(order_key, all_templates[0])
                                tmpl_key = f"{rec['廠商名稱']}_{rec['模板名稱']}"
                                wb_bytes = st.session_state.template_wb_bytes.get(tmpl_key)
                                twb = (
                                    openpyxl.load_workbook(BytesIO(wb_bytes))
                                    if wb_bytes else openpyxl.Workbook()
                                )
                                pairs.append({
                                    "order": o,
                                    "template_info": template_from_json(rec["設定JSON"]),
                                    "template_wb": twb,
                                    "template_bytes": wb_bytes,
                                    "vendor": rec.get("廠商名稱", ""),
                                })

                            buf = generate_labels_multiorder(pairs)
                            st.download_button(
                                "⬇️ 下載標籤.xlsx",
                                data=buf,
                                file_name="出貨標籤.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True,
                            )
                        except Exception as e:
                            st.error(f"產出失敗：{e}")
                            import traceback
                            st.code(traceback.format_exc())

                # 若模板 Excel 尚未上傳，提供重新上傳入口
                if selected_orders and order_tmpl_map:
                    _shown_keys = set()
                    for _ok, _rec in order_tmpl_map.items():
                        _tk = f"{_rec['廠商名稱']}_{_rec['模板名稱']}"
                        if _tk in _shown_keys or _tk in st.session_state.template_wb_bytes:
                            continue
                        _shown_keys.add(_tk)
                        _re = st.file_uploader(
                            f"上傳「{_tmpl_label(_rec)}」原始 Excel（保留樣式用）",
                            type=["xlsx", "xls"],
                            key=f"reupload_{_tk}",
                        )
                        if _re:
                            st.session_state.template_wb_bytes[_tk] = _re.read()
                            st.rerun()

    # ── 管理模板 ──────────────────────────────────────────────
    with sub_tab_manage:
        st.markdown("### 上傳新模板")
        st.caption("上傳舊的標籤 Excel，系統自動分析格式，讓你確認欄位對應後存檔；廠商名稱／模板名稱留空時，直接使用工作表名稱命名")

        col1, col2 = st.columns(2)
        with col1:
            new_customer = st.text_input("廠商名稱（留空則用工作表名稱）", placeholder="例：晶晟")
        with col2:
            new_tmpl_name = st.text_input("模板名稱（留空則用工作表名稱）", placeholder="例：標準出貨標籤")

        uploaded_tmpl = st.file_uploader(
            "上傳標籤 Excel 範本",
            type=["xlsx", "xls"],
            key="new_template_upload",
        )

        if uploaded_tmpl:
            tmpl_bytes = uploaded_tmpl.read()
            wb = openpyxl.load_workbook(BytesIO(tmpl_bytes))
            sheet_names = wb.sheetnames

            batch_mode = st.checkbox(
                "批次建立（每個工作表自動建立一個模板）",
                value=len(sheet_names) > 1,
            )

            if batch_mode:
                if st.button("批次分析所有工作表", type="primary"):
                    with st.spinner(f"分析 {len(sheet_names)} 個工作表..."):
                        results = analyze_all_sheets(wb)
                    if not results:
                        st.error("無法分析任何工作表")
                    else:
                        saved, skipped = 0, 0
                        sync_errors = []
                        first_sync_err = None
                        for sname, info in results.items():
                            try:
                                vendor = new_customer.strip() or sname
                                tname = new_tmpl_name.strip() or sname
                                _err = save_template(vendor, tname, template_to_json(info), excel_bytes=tmpl_bytes)
                                if _err:
                                    sync_errors.append(f"{vendor} — {tname}")
                                    first_sync_err = first_sync_err or _err
                                tmpl_key = f"{vendor}_{tname}"
                                st.session_state.template_wb_bytes[tmpl_key] = tmpl_bytes
                                saved += 1
                            except Exception:
                                skipped += 1
                        clear_cache()
                        st.success(f"完成！成功建立 {saved} 個模板" + (f"，{skipped} 個失敗" if skipped else ""))
                        if sync_errors:
                            st.warning(
                                "⚠️ 以下模板的原始 Excel 未能同步到雲端（僅存在目前分頁的暫存中）："
                                + "、".join(sync_errors)
                                + "。下次重新整理或換裝置時會需要重新上傳。"
                                f"\n\n錯誤訊息：{first_sync_err}"
                            )
                        else:
                            st.rerun()
            else:
                selected_sheet = st.selectbox("選擇要分析的工作表", sheet_names)
                if st.button("分析模板", type="primary"):
                    with st.spinner("分析中..."):
                        template_info = analyze_template(wb, selected_sheet)
                    if not template_info:
                        st.error("無法分析此工作表，請確認內容不為空")
                    else:
                        st.success(
                            f"分析完成：{template_info['unit_rows']} 行/標籤，"
                            f"每列 {template_info['units_per_row']} 個並排"
                        )
                        st.session_state["_pending_template"] = template_info
                        st.session_state["_pending_tmpl_bytes"] = tmpl_bytes
                        st.session_state["_pending_customer"] = new_customer.strip() or selected_sheet
                        st.session_state["_pending_tmpl_name"] = new_tmpl_name.strip() or selected_sheet
                        st.session_state["_pending_is_edit"] = False

        # 欄位對應確認
        if "_pending_template" in st.session_state:
            template_info = st.session_state["_pending_template"]
            cells = template_info.get("cells", [])

            st.divider()
            st.markdown("### 確認欄位對應")
            st.caption("檢查每個格子的欄位是否正確，可以修改")

            # 名稱可編輯（新增時預填，編輯時可改名）。
            # widget key 帶入原始名稱，確保切換到編輯「另一個」模板時輸入框會重置成新模板的名稱，
            # 不會沿用上一個模板編輯時殘留的文字（否則存檔時會誤判成新模板，另外新增一列）。
            _pending_key_suffix = (
                f"{st.session_state.get('_pending_customer','')}_{st.session_state.get('_pending_tmpl_name','')}"
            )
            _ec1, _ec2 = st.columns(2)
            with _ec1:
                _edit_customer = st.text_input(
                    "廠商名稱（用於比對）",
                    value=st.session_state.get("_pending_customer", ""),
                    key=f"_edit_customer_input_{_pending_key_suffix}",
                )
            with _ec2:
                _edit_tmpl_name = st.text_input(
                    "模板名稱",
                    value=st.session_state.get("_pending_tmpl_name", ""),
                    key=f"_edit_tmpl_name_input_{_pending_key_suffix}",
                )

            # 標籤列數（自動偵測，可手動修正）
            _auto_unit_rows = template_info.get("unit_rows", 1)
            _ec_rows, _ec_cols = st.columns(2)
            with _ec_rows:
                _edit_unit_rows = st.number_input(
                    f"每個標籤佔幾列（自動偵測={_auto_unit_rows}，模板有多份樣本時請填單份列數）",
                    min_value=1,
                    value=_auto_unit_rows,
                    step=1,
                    key="_edit_unit_rows_input",
                )
            with _ec_cols:
                st.metric("並排數（自動）", template_info.get("units_per_row", 1))

            field_options = ["__fixed__（固定文字）"] + [
                f"{k}（{v}）" for k, v in DYNAMIC_FIELDS.items() if k != "固定文字"
            ]

            updated_cells = []
            for i, cell in enumerate(cells):
                c1, c2, c3 = st.columns([1, 3, 3])
                with c1:
                    st.text(f"R{cell['row']}C{cell['col']}")
                with c2:
                    st.text(cell['value'][:40] if cell['value'] else "")
                with c3:
                    current = cell.get("field", "__fixed__")
                    # _inline 變體 (e.g. item_no_inline) → 找基底 key 對應選單
                    _cur_base = current[:-len("_inline")] if current.endswith("_inline") else current
                    # 選項格式：「__fixed__（固定文字）」或「中文名（field_key）」
                    if _cur_base == "__fixed__":
                        curr_display = next(
                            (o for o in field_options if o.startswith("__fixed__")),
                            field_options[0]
                        )
                    else:
                        curr_display = next(
                            (o for o in field_options if f"（{_cur_base}）" in o),
                            field_options[0]
                        )
                    chosen = st.selectbox(
                        "欄位",
                        field_options,
                        index=field_options.index(curr_display) if curr_display in field_options else 0,
                        key=f"field_map_{i}",
                        label_visibility="collapsed",
                    )
                    # 取出內部 field key（括號內），保留 _inline 格式
                    if chosen.startswith("__fixed__"):
                        new_field = "__fixed__"
                    else:
                        _fm = re.search(r'（([^）]+)）', chosen)
                        _base_new = _fm.group(1) if _fm else "__fixed__"
                        # 若原本 inline 且選同一欄位 → 保留 inline；否則看 cell value 有無冒號前綴
                        if current.endswith("_inline") and _base_new == _cur_base:
                            new_field = current
                        elif re.search(r'[：:]\s*$', cell.get("value", "")):
                            new_field = f"{_base_new}_inline"
                        else:
                            new_field = _base_new
                    updated_cell = {**cell, "field": new_field}
                    updated_cells.append(updated_cell)

            if st.button("儲存模板", type="primary", use_container_width=True):
                template_info["cells"] = updated_cells
                template_info["unit_rows"] = int(_edit_unit_rows)
                config_json = template_to_json(template_info)
                customer = _edit_customer.strip() or st.session_state.get("_pending_customer", "")
                tmpl_name = _edit_tmpl_name.strip() or st.session_state.get("_pending_tmpl_name", "")
                tmpl_key = f"{customer}_{tmpl_name}"
                _is_edit = st.session_state.get("_pending_is_edit", False)
                _orig_customer = st.session_state.get("_pending_customer", "") if _is_edit else None
                _orig_tmpl_name = st.session_state.get("_pending_tmpl_name", "") if _is_edit else None

                try:
                    _wb_bytes = st.session_state.get("_pending_tmpl_bytes") or b""
                    _sync_err = save_template(
                        customer, tmpl_name, config_json, excel_bytes=_wb_bytes or None,
                        original_customer=_orig_customer, original_template_name=_orig_tmpl_name,
                    )
                    # 快取 workbook bytes
                    st.session_state.template_wb_bytes[tmpl_key] = _wb_bytes
                    # 清除暫存
                    del st.session_state["_pending_template"]
                    del st.session_state["_pending_tmpl_bytes"]
                    del st.session_state["_pending_customer"]
                    del st.session_state["_pending_tmpl_name"]
                    st.session_state.pop("_pending_is_edit", None)
                    clear_cache()
                    st.success(f"模板「{customer} — {tmpl_name}」已儲存！")
                    if _sync_err:
                        st.warning(
                            f"⚠️ 原始 Excel 未能同步到雲端，重新整理頁面或換裝置後會需要重新上傳。\n\n"
                            f"錯誤訊息：{_sync_err}"
                        )
                    else:
                        st.rerun()
                except Exception as e:
                    st.error(f"儲存失敗：{e}")

        # 現有模板列表
        st.divider()
        st.markdown("### 現有模板")
        try:
            all_templates = load_templates()
            if not all_templates:
                st.info("尚無模板")
            else:
                for r in all_templates:
                    with st.expander(f"{_tmpl_label(r)}　（更新：{r.get('最後更新','')}）"):
                        info = template_from_json(r["設定JSON"])
                        st.write(f"標籤行數：{info.get('unit_rows')}，並排數：{info.get('units_per_row')}")
                        cells = info.get("cells", [])
                        dynamic = [c for c in cells if c.get("field") != "__fixed__"]
                        st.write(f"動態欄位：{[c['field'] for c in dynamic]}")
                        if not dynamic:
                            st.warning("⚠️ 無動態欄位，標籤將全為固定文字！請重新上傳 Excel 分析，或手動編輯欄位。")
                        _btn_edit, _btn_del = st.columns(2)
                        with _btn_edit:
                            if st.button("✏️ 編輯欄位", key=f"edit_{r['廠商名稱']}_{r['模板名稱']}"):
                                st.session_state["_pending_template"] = dict(info)
                                st.session_state["_pending_tmpl_bytes"] = st.session_state.template_wb_bytes.get(
                                    f"{r['廠商名稱']}_{r['模板名稱']}", b""
                                )
                                st.session_state["_pending_customer"] = r["廠商名稱"]
                                st.session_state["_pending_tmpl_name"] = r["模板名稱"]
                                st.session_state["_pending_is_edit"] = True
                                st.rerun()
                        with _btn_del:
                            if st.button("🗑 刪除此模板", key=f"del_{r['廠商名稱']}_{r['模板名稱']}"):
                                try:
                                    delete_template(r["廠商名稱"], r["模板名稱"])
                                    clear_cache()
                                    st.success("已刪除")
                                    st.rerun()
                                except Exception as e:
                                    st.error(f"刪除失敗：{e}")
                        # 重新分析（上傳 Excel → 覆寫欄位設定）
                        _tmpl_key = f"{r['廠商名稱']}_{r['模板名稱']}"
                        _re_file = st.file_uploader(
                            "🔄 重新分析（上傳原始 Excel 覆寫欄位）",
                            type=["xlsx", "xls"],
                            key=f"reanalyze_{_tmpl_key}",
                        )
                        if _re_file:
                            _re_bytes = _re_file.read()
                            _re_wb = openpyxl.load_workbook(BytesIO(_re_bytes))
                            _re_sname = info.get("sheet_name", _re_wb.sheetnames[0])
                            if _re_sname not in _re_wb.sheetnames:
                                _re_sname = _re_wb.sheetnames[0]
                            _re_info = analyze_template(_re_wb, _re_sname)
                            if _re_info and _re_info.get("cells"):
                                _re_err = save_template(r["廠商名稱"], r["模板名稱"], template_to_json(_re_info), excel_bytes=_re_bytes)
                                st.session_state.template_wb_bytes[_tmpl_key] = _re_bytes
                                clear_cache()
                                st.success(f"重新分析完成，找到 {len([c for c in _re_info['cells'] if c['field']!='__fixed__'])} 個動態欄位")
                                if _re_err:
                                    st.warning(
                                        f"⚠️ 原始 Excel 未能同步到雲端，下次重新整理可能又要重傳一次。\n\n"
                                        f"錯誤訊息：{_re_err}"
                                    )
                                else:
                                    st.rerun()
                            else:
                                st.error("分析失敗，此工作表無內容")
        except Exception as e:
            st.error(f"載入失敗：{e}")

    # ── 從廠商網站下載標籤 ───────────────────────────────────────
    with sub_tab_erp:
        st.caption("從廠商 ERP 網站下載標籤 PDF，可從已解析的銷貨單勾選或直接輸入單號")

        # ── 廠商帳號管理 ──────────────────────────────────────────
        _BUILTIN_VENDORS = [
            {"公司名稱": "鴻勁", "網址": "http://scm.honprec.com/hp/Index.aspx",
             "帳號": "BR026", "密碼": "5403"},
        ]

        try:
            _custom_vendors = load_vendors()
        except Exception:
            _custom_vendors = []

        # 合併：內建優先，避免重複
        _builtin_names = {v["公司名稱"] for v in _BUILTIN_VENDORS}
        _all_vendors = _BUILTIN_VENDORS + [v for v in _custom_vendors if v.get("公司名稱") not in _builtin_names]

        if "vendor_selected" not in st.session_state:
            st.session_state.vendor_selected = ""
        if "show_add_vendor" not in st.session_state:
            st.session_state.show_add_vendor = False

        st.markdown("**選擇廠商：**")
        _vcols = st.columns(len(_all_vendors) + 1)
        for _vi, _v in enumerate(_all_vendors):
            with _vcols[_vi]:
                _vtype = "primary" if st.session_state.vendor_selected == _v["公司名稱"] else "secondary"
                if st.button(_v["公司名稱"], key=f"vbtn_{_vi}", type=_vtype, use_container_width=True):
                    st.session_state.vendor_selected = _v["公司名稱"]
                    st.session_state.show_add_vendor = False
                    st.rerun()
        with _vcols[-1]:
            if st.button("＋ 新增廠商", key="add_vendor_btn", use_container_width=True):
                st.session_state.show_add_vendor = not st.session_state.show_add_vendor
                st.rerun()

        # 新增廠商表單
        if st.session_state.show_add_vendor:
            with st.form("add_vendor_form"):
                _fc1, _fc2 = st.columns(2)
                with _fc1:
                    _vname = st.text_input("公司名稱 *")
                    _vurl  = st.text_input("ERP 登入網址 *")
                with _fc2:
                    _vuser = st.text_input("帳號")
                    _vpass = st.text_input("密碼", type="password")
                _fs, _fc = st.columns(2)
                with _fs:
                    _do_save = st.form_submit_button("儲存", type="primary")
                with _fc:
                    _do_cancel = st.form_submit_button("取消")

                if _do_save:
                    if _vname and _vurl:
                        try:
                            save_vendor(_vname, _vurl, _vuser, _vpass)
                            clear_cache()
                            st.session_state.show_add_vendor = False
                            st.session_state.vendor_selected = _vname
                            st.rerun()
                        except Exception as _ve:
                            st.error(f"儲存失敗：{_ve}")
                    else:
                        st.warning("公司名稱和網址為必填")
                if _do_cancel:
                    st.session_state.show_add_vendor = False
                    st.rerun()

        # ── 已選廠商 → 下載介面 ──────────────────────────────────
        _sel_v = st.session_state.vendor_selected
        if _sel_v:
            _vrec = next((v for v in _all_vendors if v["公司名稱"] == _sel_v), None)
            if _vrec:
                # 非內建廠商才顯示刪除
                if _sel_v not in _builtin_names:
                    with st.expander(f"「{_sel_v}」設定"):
                        st.caption(f"網址：{_vrec.get('網址','')}")
                        if st.button(f"🗑 刪除「{_sel_v}」", key="del_vendor_btn"):
                            try:
                                delete_vendor(_sel_v)
                                clear_cache()
                                st.session_state.vendor_selected = ""
                                st.rerun()
                            except Exception as _de:
                                st.error(f"刪除失敗：{_de}")
                else:
                    st.info(f"已選擇：**{_sel_v}**　　{_vrec.get('網址','')}")

                _erp_all_orders = st.session_state.parsed_orders
                _order_nos_parsed = [o.get("order_no", "") for o in _erp_all_orders if o.get("order_no")]

                _input_mode = st.radio(
                    "選擇出貨單方式",
                    ["從已解析銷貨單勾選", "直接輸入銷貨單號"],
                    horizontal=True,
                    key="erp_input_mode",
                )

                _selected_nos: list[str] = []
                if _input_mode == "從已解析銷貨單勾選":
                    if not _order_nos_parsed:
                        st.info("尚無已解析的銷貨單，請先在左側上傳，或改用「直接輸入銷貨單號」")
                    else:
                        _selected_nos = st.multiselect(
                            "選擇要下載的出貨單",
                            options=_order_nos_parsed,
                            default=_order_nos_parsed,
                            key="erp_order_multiselect",
                        )
                else:
                    _raw_text = st.text_area(
                        "輸入銷貨單號（可多筆，用換行或逗號分隔）",
                        placeholder="202606100012\n202606020007",
                        height=100,
                        key="erp_order_text",
                    )
                    _selected_nos = [n.strip() for n in re.split(r"[,\n，、]+", _raw_text) if n.strip()]
                    if _selected_nos:
                        st.caption(f"共 {len(_selected_nos)} 筆")

                if _selected_nos and st.button("從廠商網站下載標籤", type="primary",
                                               use_container_width=True, key="erp_download_btn"):
                    with st.spinner("連線 ERP 並下載中，請稍候..."):
                        try:
                            from erp_downloader import download_label_pdfs, pack_zip
                            _results, _erp_errors = download_label_pdfs(
                                _selected_nos,
                                _vrec.get("帳號", ""),
                                _vrec.get("密碼", ""),
                            )
                            _success = {no: d for no, d in _results.items() if d}
                            _failed  = [no for no, d in _results.items() if not d]

                            def _pdf_to_combined_b64(pdf_bytes: bytes, dpi: int = 150) -> str:
                                """把 PDF 所有頁面垂直合併成一張 PNG，回傳 base64"""
                                import fitz as _fitz
                                import base64 as _b64
                                from PIL import Image as _PILImage
                                _doc = _fitz.open(stream=pdf_bytes, filetype="pdf")
                                _imgs = [
                                    _PILImage.frombytes(
                                        "RGB",
                                        [_p.width, _p.height],
                                        _p.samples,
                                    )
                                    for _p in (_doc[i].get_pixmap(dpi=dpi) for i in range(len(_doc)))
                                ]
                                _w = max(im.width for im in _imgs)
                                _h = sum(im.height for im in _imgs)
                                _combined = _PILImage.new("RGB", (_w, _h), "white")
                                _y = 0
                                for im in _imgs:
                                    _combined.paste(im, (0, _y))
                                    _y += im.height
                                _buf = BytesIO()
                                _combined.save(_buf, format="PNG")
                                return _b64.b64encode(_buf.getvalue()).decode()

                            def _copy_button_html(b64: str, btn_id: str, label: str) -> str:
                                return f"""
<button id="{btn_id}" onclick="copyLabel_{btn_id}()" style="
    background:#0068c9;color:white;border:none;border-radius:6px;
    padding:8px 0;font-size:15px;cursor:pointer;width:100%;margin-top:4px">
    {label}
</button>
<div id="toast_{btn_id}" style="display:none;margin-top:6px;padding:8px;
    background:#21c354;color:white;border-radius:6px;text-align:center;font-size:14px">
    ✓ 已複製！可直接貼上到 Codex
</div>
<script>
async function copyLabel_{btn_id}(){{
    try{{
        const blob=await fetch('data:image/png;base64,{b64}').then(r=>r.blob());
        await navigator.clipboard.write([new ClipboardItem({{'image/png':blob}})]);
        document.getElementById('toast_{btn_id}').style.display='block';
        document.getElementById('{btn_id}').textContent='✓ 已複製';
        document.getElementById('{btn_id}').style.background='#21c354';
        setTimeout(()=>{{
            document.getElementById('toast_{btn_id}').style.display='none';
            document.getElementById('{btn_id}').textContent='{label}';
            document.getElementById('{btn_id}').style.background='#0068c9';
        }},2500);
    }}catch(e){{alert('複製失敗：'+e.message);}}
}}
</script>"""

                            if _success:
                                if len(_success) == 1:
                                    _ono, _pdf = next(iter(_success.items()))
                                    st.download_button(
                                        f"⬇️ 下載 {_ono} 標籤.pdf",
                                        data=_pdf,
                                        file_name=f"標籤_{_ono}.pdf",
                                        mime="application/pdf",
                                    )
                                    try:
                                        _b64str = _pdf_to_combined_b64(_pdf)
                                        st.components.v1.html(
                                            _copy_button_html(_b64str, "cp_single", "複製截圖"),
                                            height=100,
                                        )
                                        st.image(
                                            BytesIO(__import__('base64').b64decode(_b64str)),
                                            caption=f"標籤預覽（全頁）：{_ono}",
                                            use_container_width=True,
                                        )
                                    except Exception as _pe:
                                        st.warning(f"無法產生預覽：{_pe}")
                                else:
                                    _zb = pack_zip(_success)
                                    st.download_button(
                                        f"⬇️ 下載所有標籤 ({len(_success)} 筆).zip",
                                        data=_zb,
                                        file_name="ERP標籤.zip",
                                        mime="application/zip",
                                        use_container_width=True,
                                    )
                                    try:
                                        for _mno, _mpdf in _success.items():
                                            _mb64 = _pdf_to_combined_b64(_mpdf)
                                            _bid = f"cp_{_mno.replace('-','_')}"
                                            st.components.v1.html(
                                                _copy_button_html(_mb64, _bid, f"複製截圖（{_mno}）"),
                                                height=100,
                                            )
                                            st.image(
                                                BytesIO(__import__('base64').b64decode(_mb64)),
                                                caption=f"標籤預覽（全頁）：{_mno}",
                                                use_container_width=True,
                                            )
                                    except Exception as _mpe:
                                        st.warning(f"無法產生預覽：{_mpe}")

                            if _failed:
                                st.warning(f"以下出貨單下載失敗：{', '.join(_failed)}")
                                for _fn in _failed:
                                    if _fn in _erp_errors:
                                        st.code(f"[{_fn}] {_erp_errors[_fn]}")
                            if not _success and not _failed:
                                st.error("未取得任何 PDF，請確認帳密與網路連線")

                        except ImportError as _ie:
                            st.error(f"缺少套件：{_ie}")
                        except Exception as _ee:
                            import traceback as _etb
                            st.error(f"ERP 下載失敗：{_ee}")
                            st.code(_etb.format_exc())


    # ── LSCR 確認單直接產出 ────────────────────────────────────────
    with sub_tab_lscr:
        st.caption("上傳 LSCR 出貨明細確認單（xlsx），自動解析明細並用內建 lable 工作表產出標籤"
                   "（若檔案只有 list 沒有 lable，會自動沿用雲端的預設模板0）")

        _lscr_up = st.file_uploader(
            "上傳 LSCR 確認單 xlsx",
            type=["xlsx"],
            key="lscr_up",
        )

        if _lscr_up:
            _lscr_bytes = _lscr_up.read()
            try:
                _wb_data = openpyxl.load_workbook(BytesIO(_lscr_bytes), data_only=True)

                if "list" not in _wb_data.sheetnames:
                    st.error("此檔案缺少工作表：list")
                    _tmpl_bytes = None
                elif "lable" in _wb_data.sheetnames:
                    # 這次上傳的檔案自帶 lable：直接用它，並在雲端還沒有模板0 時存一份
                    # （只存第一次，之後不再覆蓋，避免之後上傳的檔案版型跑掉時把模板0 也帶壞）
                    _tmpl_bytes = _lscr_bytes
                    try:
                        if not find_lscr_base_template_id():
                            save_lscr_base_template(_lscr_bytes)
                            download_lscr_base_template.clear()
                            st.info("已將本次的 lable 版型另存為雲端預設模板0，"
                                    "之後上傳只有 list 的檔案會自動沿用。")
                    except Exception as _sav_e:
                        st.warning(f"雲端模板0 儲存失敗（不影響本次產出）：{_sav_e}")
                else:
                    # 檔案只有 list 沒有 lable：沿用雲端的預設模板0
                    _base_bytes = download_lscr_base_template()
                    if _base_bytes:
                        _tmpl_bytes = _base_bytes
                        st.caption("此檔案沒有 lable 工作表，已自動沿用雲端預設模板0 排版")
                    else:
                        st.error("此檔案缺少工作表：lable，且雲端尚無預設模板0 可沿用"
                                 "（請先上傳一份含 lable 工作表的檔案，之後才能沿用）")
                        _tmpl_bytes = None

                if _tmpl_bytes:
                    _wb_tmpl = openpyxl.load_workbook(BytesIO(_tmpl_bytes))
                    _lscr_orders = parse_lscr_excel_wb(_wb_data)
                    _lscr_raw_items = [
                        (itm, o)
                        for o in _lscr_orders
                        for itm in o.get("items", [])
                    ]

                    st.success(f"解析完成：{len(_lscr_orders)} 張訂單，共 {len(_lscr_raw_items)} 個品項")

                    # 品項預覽
                    st.dataframe(
                        [{
                            "PO NO":    o.get("order_no", ""),
                            "料號":      itm.get("item_no", ""),
                            "品名":      itm.get("name", ""),
                            "規格":      itm.get("description", ""),
                            "總數量":    f"{itm.get('_total_qty','')}{itm.get('_large_unit','PCS')}",
                            "大包裝":    f"{itm.get('_large_qty','')}{itm.get('_large_unit','PCS')}",
                            "小包裝":    f"{itm.get('_small_qty','')}{itm.get('_small_unit','PCS')}",
                            "LOT NO":   itm.get("lot_no", ""),
                        } for itm, o in _lscr_raw_items],
                        use_container_width=True,
                        hide_index=True,
                        height=min(420, 38 * (len(_lscr_raw_items) + 1) + 10),
                        column_config={
                            "PO NO":  st.column_config.TextColumn(width="medium"),
                            "料號":    st.column_config.TextColumn(width="large"),
                            "品名":    st.column_config.TextColumn(width="medium"),
                            "規格":    st.column_config.TextColumn(width="large"),
                            "總數量":  st.column_config.TextColumn(width="small"),
                            "大包裝":  st.column_config.TextColumn(width="small"),
                            "小包裝":  st.column_config.TextColumn(width="small"),
                            "LOT NO": st.column_config.TextColumn(width="medium"),
                        },
                    )

                    # 分裝設定
                    st.markdown("**標籤設定**")
                    _lscr_c1, _lscr_c2 = st.columns(2)
                    with _lscr_c1:
                        _lscr_small = st.checkbox("印小標籤（小包裝數量）", value=True, key="lscr_small")
                    with _lscr_c2:
                        _lscr_large = st.checkbox("印大標籤（總出貨數量）", value=True, key="lscr_large")

                    if st.button("產出標籤 Excel", type="primary",
                                 use_container_width=True, key="lscr_gen"):
                        with st.spinner("產出中..."):
                            try:
                                _lscr_tmpl_info = analyze_template(_wb_tmpl, "lable")
                                if not _lscr_tmpl_info or not _lscr_tmpl_info.get("cells"):
                                    st.error("無法分析 lable 工作表")
                                else:
                                    _buf = write_lscr_labels(
                                        _lscr_orders,
                                        openpyxl.load_workbook(BytesIO(_tmpl_bytes)),
                                        _lscr_tmpl_info,
                                        include_small=_lscr_small,
                                        include_large=_lscr_large,
                                        tmpl_bytes=_tmpl_bytes,
                                    )
                                    st.download_button(
                                        "⬇️ 下載標籤.xlsx",
                                        data=_buf,
                                        file_name="LSCR標籤.xlsx",
                                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                        use_container_width=True,
                                    )
                            except Exception as _le:
                                import traceback as _ltb
                                st.error(f"產出失敗：{_le}")
                                st.code(_ltb.format_exc())

            except Exception as _lscr_e:
                st.error(f"解析失敗：{_lscr_e}")
                import traceback as _ltb2
                st.code(_ltb2.format_exc())


# ════════════════════════════════════════════════════════════════
#  電子發票
# ════════════════════════════════════════════════════════════════
with tab_invoice:
    st.subheader("電子發票")
    st.caption("依照 e-invoice.com.tw V1.6 格式產生上傳檔")

    _inv_tab_rtf, _inv_tab_monthly = st.tabs(["從銷貨單", "月結（驗收資訊）"])

    # ── 從銷貨單 ─────────────────────────────────────────────────
    with _inv_tab_rtf:
        all_orders = st.session_state.parsed_orders
        if not all_orders:
            st.info("請先在左側上傳並解析銷貨單")
        else:
            order_map = {
                f"{o.get('order_no','')} — {o.get('customer_code','')} ({len(o.get('items',[]))} 項)": o
                for o in all_orders
            }
            selected_keys = st.multiselect(
                "選擇要開發票的銷貨單",
                options=list(order_map.keys()),
                default=list(order_map.keys()),
            )
            selected_orders = [order_map[k] for k in selected_keys if order_map[k].get("items")]

            if selected_orders:
                _with_inv    = [o for o in selected_orders if o.get("invoice_no")]
                _without_inv = [o for o in selected_orders if not o.get("invoice_no")]

                if _with_inv:
                    st.success(f"共 **{len(_with_inv)}** 張有發票號碼，將產出發票")
                if _without_inv:
                    _no_inv_nos = [o.get("order_no","(未知)") for o in _without_inv]
                    st.warning(f"{len(_without_inv)} 張無發票號碼（略過）：{', '.join(_no_inv_nos)}")

                st.caption("發票人統編：**24405403**　受票人統編：從銷貨單統一編號　單價：暫定 100（待確認）")

                if _with_inv and st.button("產出電子發票 xls", type="primary",
                                           use_container_width=True, key="inv_rtf_gen"):
                    try:
                        buf = generate_invoice_excel(_with_inv)
                        st.download_button(
                            "⬇️ 下載電子發票上傳檔.xls",
                            data=buf,
                            file_name="電子發票上傳.xls",
                            mime="application/vnd.ms-excel",
                            use_container_width=True,
                        )
                    except Exception as _inv_e:
                        import traceback as _inv_tb
                        st.error(f"產出失敗：{_inv_e}")
                        st.code(_inv_tb.format_exc())
            else:
                st.warning("請選擇至少一張銷貨單")

    # ── 月結（驗收資訊 xlsx） ─────────────────────────────────────
    with _inv_tab_monthly:
        st.caption("上傳從 ERP 匯出的「驗收資訊」xlsx，整批開一張月結發票")

        _acc_up = st.file_uploader(
            "上傳驗收資訊 xlsx",
            type=["xlsx"],
            key="acc_up",
        )

        if _acc_up:
            _acc_bytes = _acc_up.read()
            try:
                _acc_rows = parse_acceptance_excel(_acc_bytes)
            except Exception as _ae:
                st.error(f"解析失敗：{_ae}")
                _acc_rows = []

            if _acc_rows:
                _acc_total = sum(r["amount"] for r in _acc_rows)
                st.success(f"解析完成：**{len(_acc_rows)}** 行，未稅合計 **{_acc_total:,.0f}**，"
                           f"稅額 **{round(_acc_total*0.05):,.0f}**，"
                           f"含稅 **{round(_acc_total*1.05):,.0f}**")

                # 品項預覽
                st.dataframe(
                    [{
                        "出貨單號": r["order_no"],
                        "單號":     r["line_no"],
                        "品號":     r["part_no"],
                        "品名":     r["name"],
                        "規格":     r["spec"],
                        "數量":     r["qty"],
                        "單價":     r["unit_price"],
                        "金額(未稅)": r["amount"],
                    } for r in _acc_rows],
                    use_container_width=True,
                    hide_index=True,
                    height=min(400, 38 * (len(_acc_rows) + 1) + 10),
                )

                st.markdown("**發票資訊**")
                _inv_col1, _inv_col2, _inv_col3 = st.columns(3)
                with _inv_col1:
                    _acc_inv_no = st.text_input("發票號碼", placeholder="AA00000000", key="acc_inv_no")
                with _inv_col2:
                    _acc_inv_date = st.date_input("發票日期", key="acc_inv_date")
                with _inv_col3:
                    _acc_buyer_id = st.text_input("受票人統編", placeholder="12345678", key="acc_buyer_id")

                st.caption("發票人統編：**24405403**")

                _can_gen = bool(_acc_inv_no and _acc_buyer_id)
                if not _can_gen:
                    st.warning("請填入發票號碼與受票人統編")

                if _can_gen and st.button("產出電子發票 xls", type="primary",
                                          use_container_width=True, key="acc_gen"):
                    try:
                        _acc_date_str = _acc_inv_date.strftime("%Y%m%d")
                        _acc_buf = generate_invoice_from_acceptance(
                            _acc_rows,
                            invoice_no=_acc_inv_no.strip(),
                            invoice_date=_acc_date_str,
                            buyer_tax_id=_acc_buyer_id.strip(),
                        )
                        st.download_button(
                            "⬇️ 下載電子發票上傳檔.xls",
                            data=_acc_buf,
                            file_name=f"電子發票_{_acc_inv_no.strip()}.xls",
                            mime="application/vnd.ms-excel",
                            use_container_width=True,
                        )
                    except Exception as _ae2:
                        import traceback as _atb
                        st.error(f"產出失敗：{_ae2}")
                        st.code(_atb.format_exc())