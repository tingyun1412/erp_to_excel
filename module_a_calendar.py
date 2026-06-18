"""
模組 A：銷貨單 → 出貨行事曆
將多張銷貨單的出貨資訊填入倉庫出貨行事曆（.xlsb 格式的欄位結構）
輸出為 .xlsx（openpyxl 支援）
"""
import calendar
from datetime import datetime, date
from io import BytesIO

import openpyxl
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side
)


# ── 樣式常數 ──────────────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT   = Font(color="FFFFFF", bold=True, size=10)
WEEKDAY_FILLS = {
    0: PatternFill("solid", fgColor="D9E1F2"),  # 一
    1: PatternFill("solid", fgColor="D9E1F2"),  # 二
    2: PatternFill("solid", fgColor="D9E1F2"),  # 三
    3: PatternFill("solid", fgColor="D9E1F2"),  # 四
    4: PatternFill("solid", fgColor="D9E1F2"),  # 五
    5: PatternFill("solid", fgColor="FFE699"),  # 六
    6: PatternFill("solid", fgColor="FFE699"),  # 日
}
ITEM_FILL     = PatternFill("solid", fgColor="FFFFFF")
ALT_FILL      = PatternFill("solid", fgColor="F2F2F2")
THIN          = Side(style="thin", color="BFBFBF")
BORDER        = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER        = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT          = Alignment(horizontal="left", vertical="center", wrap_text=True)

WEEKDAY_ZH = ["一", "二", "三", "四", "五", "六", "日"]


def _style(cell, fill=None, font=None, alignment=None, border=None):
    if fill:      cell.fill      = fill
    if font:      cell.font      = font
    if alignment: cell.alignment = alignment
    if border:    cell.border    = border


def _tw_year_month(year: int, month: int) -> str:
    """西元年月轉民國年月，例：2026/9 → 11509"""
    tw_year = year - 1911
    return f"{tw_year:03d}{month:02d}"


def build_calendar_sheet(ws, year: int, month: int, orders: list[dict]):
    """
    在 ws 工作表上建立一個月的出貨行事曆
    
    orders: list of dict，每筆包含：
        ship_date (str YYYYMMDD), customer, item_no, description,
        quantity, remark
    """
    tw_ym = _tw_year_month(year, month)
    month_zh = f"{month}月份"
    
    # ── 標題列 ──────────────────────────────────────────────────
    ws.merge_cells("A1:B1")
    ws["A1"] = "日期"
    ws.merge_cells("C1:D1")
    ws["C1"] = "數量"
    
    title_cell = ws.cell(row=1, column=8, value="倉庫出貨通知")
    ws.merge_cells(start_row=1, start_column=8, end_row=1, end_column=12)
    _style(title_cell, fill=HEADER_FILL, font=HEADER_FONT, alignment=CENTER)
    
    ws.cell(row=1, column=14, value="客供品")
    ws.cell(row=1, column=15, value="標籤OK")
    ws.cell(row=1, column=16, value="已完成")
    
    ws.merge_cells("H2:L2")
    month_cell = ws.cell(row=2, column=8, value=month_zh)
    _style(month_cell, fill=HEADER_FILL, font=HEADER_FONT, alignment=CENTER)
    ws.cell(row=2, column=14, value="借出")
    ws.cell(row=2, column=15, value="換貨")
    ws.cell(row=2, column=16, value="Free")
    
    # ── 按出貨日期分組 ──────────────────────────────────────────
    from collections import defaultdict
    by_date = defaultdict(list)
    for order in orders:
        sd = order.get("ship_date", "")
        if len(sd) == 8:
            try:
                d = date(int(sd[:4]), int(sd[4:6]), int(sd[6:8]))
                if d.year == year and d.month == month:
                    by_date[d].append(order)
            except ValueError:
                pass
    
    # ── 計算每週的日期範圍 ──────────────────────────────────────
    cal = calendar.monthcalendar(year, month)  # 每列是一週 [Mon..Sun]
    
    current_row = 3
    
    for week_idx, week in enumerate(cal):
        # 週標題行：星期幾
        header_row = current_row
        ws.cell(row=header_row, column=1, value="日期")
        ws.cell(row=header_row, column=2, value="交期延誤")
        ws.cell(row=header_row, column=3, value="數量")
        
        col = 4
        for wd_idx, day_num in enumerate(week):
            wd_zh = WEEKDAY_ZH[wd_idx]
            cell = ws.cell(row=header_row, column=col, value=wd_zh)
            _style(cell, fill=WEEKDAY_FILLS[wd_idx], alignment=CENTER)
            ws.merge_cells(start_row=header_row, start_column=col,
                           end_row=header_row, end_column=col+1)
            col += 2
        
        current_row += 1
        
        # 日期數字行
        date_row = current_row
        col = 4
        for wd_idx, day_num in enumerate(week):
            val = day_num if day_num != 0 else ""
            cell = ws.cell(row=date_row, column=col, value=val)
            _style(cell, fill=WEEKDAY_FILLS[wd_idx], alignment=CENTER)
            # 數量欄
            qty_cell = ws.cell(row=date_row, column=col+1, value="數量" if day_num else "")
            _style(qty_cell, fill=WEEKDAY_FILLS[wd_idx], alignment=CENTER)
            col += 2
        
        current_row += 1
        
        # 找這週最多幾筆訂單（決定要畫幾行）
        max_items = 0
        for wd_idx, day_num in enumerate(week):
            if day_num == 0:
                continue
            d = date(year, month, day_num)
            max_items = max(max_items, len(by_date[d]))
        
        # 至少留 4 行空白
        rows_needed = max(max_items, 4)
        
        for row_offset in range(rows_needed):
            row = current_row + row_offset
            fill = ITEM_FILL if row_offset % 2 == 0 else ALT_FILL
            
            col = 4
            for wd_idx, day_num in enumerate(week):
                if day_num == 0:
                    ws.cell(row=row, column=col, value="")
                    ws.cell(row=row, column=col+1, value="")
                    col += 2
                    continue
                
                d = date(year, month, day_num)
                items_today = by_date[d]
                
                if row_offset < len(items_today):
                    item = items_today[row_offset]
                    # 客戶名稱 + 料號 + 品名 分行顯示
                    customer = item.get("customer", "")
                    item_no  = item.get("item_no", "")
                    desc     = item.get("description", "")
                    qty      = item.get("quantity", "")
                    remark   = item.get("remark", "")
                    
                    lines = []
                    if customer: lines.append(customer)
                    if item_no:  lines.append(item_no)
                    if desc:     lines.append(desc)
                    if remark:   lines.append(remark)
                    
                    name_cell = ws.cell(row=row, column=col, value="\n".join(lines))
                    _style(name_cell, fill=fill, alignment=LEFT, border=BORDER)
                    
                    qty_cell = ws.cell(row=row, column=col+1, value=qty if qty else "")
                    _style(qty_cell, fill=fill, alignment=CENTER, border=BORDER)
                else:
                    ws.cell(row=row, column=col, value="")
                    ws.cell(row=row, column=col+1, value="")
                
                col += 2
            
            ws.row_dimensions[row].height = 45
        
        current_row += rows_needed + 1  # 空一行分隔週
    
    # ── 欄寬 ──────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 6
    for i, col_letter in enumerate(["D","E","F","G","H","I","J","K","L","M","N","O"]):
        ws.column_dimensions[col_letter].width = 14 if i % 2 == 0 else 8


def generate_shipping_calendar(orders: list[dict]) -> BytesIO:
    """
    接收所有訂單，依年月分組產生工作表
    orders: list of dict（見 build_calendar_sheet 說明）
    回傳 BytesIO（xlsx）
    """
    from collections import defaultdict
    
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # 刪掉預設空白 sheet
    
    # 依年月分組
    by_ym = defaultdict(list)
    for order in orders:
        sd = order.get("ship_date", "")
        if len(sd) == 8:
            try:
                y, m = int(sd[:4]), int(sd[4:6])
                by_ym[(y, m)].append(order)
            except ValueError:
                pass
    
    if not by_ym:
        # 沒有日期資料，建一個當月空白
        today = datetime.today()
        by_ym[(today.year, today.month)] = []
    
    for (y, m) in sorted(by_ym.keys()):
        sheet_name = _tw_year_month(y, m)
        ws = wb.create_sheet(title=sheet_name)
        build_calendar_sheet(ws, y, m, by_ym[(y, m)])
    
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
